import time
from random import choice
from sbvirtualdisplay.unicodeutil import unidecode
import pyautogui
import openpyxl
from seleniumbase import Driver
from selenium.common.exceptions import NoSuchElementException
from unidecode import unidecode


driver = Driver(uc=True)


def get_data(site_url):
    try:
        html_path = 'D:/'
        driver.uc_open(site_url)
        if 'Just a moment...' in driver.title:
            print('Just a moment...')
            target_location = (50, 270)
            pyautogui.moveTo(target_location[0], target_location[1], duration=0.1)
            pyautogui.click()
        else:
            pass
        html = driver.page_source
        with open(f'{html_path}{row_num}.html', 'w', encoding='utf-8') as f:
            f.write(unidecode(html))
    except (Exception, NoSuchElementException) as e:
        error_message = str(e).split('\n')[0]
        print(f"Error: {error_message} {site_url}")


# Open Xls File which contains urls.
workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row


for row_num in range(1, max_row + 1):
    url = sheet.cell(row=row_num, column=1).value
    get_data(url)
    print(url)
    if row_num % 150 == 0:
        print("Waiting for 30 seconds", end=' ')
        time.sleep(30)
    print(row_num, url, end=' ')
    waiting = choice([1, 1, 2, 2, 3, 3, 4, 4, 5, 5])
    print(f"Waiting for {waiting} seconds")
    time.sleep(waiting)
