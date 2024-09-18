import openpyxl
import json
from datetime import datetime
import urllib3
import requests
from requests.exceptions import ProxyError
from lxml import html
from urllib3.exceptions import MaxRetryError

workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['urls']
max_row = sheet.max_row
json_file_path = 'SmartBuyGlasses_Sample_01.json'

# Function to append a dict to a list and save it to a JSON file
def append_dict_to_list_and_save(dictionary, file_path):
    try:
        # Read the existing data from the JSON file (if it exists)
        with open(file_path, 'r') as file:
            data = json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        # If the file doesn't exist or is empty, create an empty list
        data = []
    data.append(dictionary)
    # Save the updated data back to the JSON file
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)


def data_parser(site_url):
    headers = {'authority': 'www.smartbuyglasses.com', 'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7', 'accept-language': 'en-US,en;q=0.9', 'cache-control': 'max-age=0', 'sec-ch-ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"', 'sec-ch-ua-mobile': '?0', 'sec-ch-ua-platform': '"Windows"', 'sec-fetch-dest': 'document', 'sec-fetch-mode': 'navigate', 'sec-fetch-site': 'none', 'sec-fetch-user': '?1', 'upgrade-insecure-requests': '1', 'user-agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"}
    if "contact-lenses" not in site_url:
        response = requests.get(site_url, headers=headers, timeout=10)

        if response.status_code == 200:
                tree = html.fromstring(response.content)
                script_tag = tree.xpath("//script[@type='application/ld+json' and contains(.,'sku')]")
                if script_tag:
                    parsed_data = {}
                    json_object = script_tag[0].text
                    # Remove invalid escape characters (replace \ with \\)
                    cleaned_json_object = json_object.replace("\\", "\\\\")
                    # Now attempt to parse the cleaned JSON string
                    product_data = json.loads(cleaned_json_object, strict=False)
                    # Now you can extract data using the product_data dictionary
                    name = product_data.get('name')
                    sku = product_data.get('sku')
                    brand = product_data.get('Brand')
                    image_urls = []
                    image = product_data.get('image')
                    image_urls.append(image)
                    category = tree.xpath("//ul[@class='top-directory-nav mb-4 pb-3']/li[2]/a/text()")[0]
                    sub_category = tree.xpath("//ul[@class='top-directory-nav mb-4 pb-3']/li[3]/a/text()")[0]
                    description = product_data.get('description')
                    try:
                        size = tree.xpath(
                            "//li[@data-single-buy-button-label='Choose Lenses' and @class='d-inline-block active']/text()")[
                            0]
                    except IndexError:
                        size = ''

                    price_1 = tree.xpath("//span[@class='ms-2 origin-price text-decoration-line-through ']/text()")
                    price_2 = tree.xpath("//span[@id='discount_price_promotion_display']/text()")

                    if len(price_1) != 0 and len(price_2) != 0:
                            regular_price = price_1[0].split('/')[0].replace(' ', '').replace('$', '').replace('\n', '')
                            discount_price = price_2[0].split('/')[0].replace(' ', '').replace('$', '').replace('\n', '')
                    elif len(price_1) != 0 and len(price_2) == 0:
                        regular_price = price_1[0].split('/')[0].replace(' ', '').replace('$', '').replace('\n', '')
                        discount_price = None
                    elif len(price_1) == 0 and len(price_2) != 0:
                        regular_price = price_2[0].split('/')[0].replace(' ', '').replace('$', '').replace('\n', '')
                        discount_price = None
                    else:
                        regular_price = None
                        discount_price = None
                    priceCurrency = product_data.get('offers', {}).get('priceCurrency', '')
                    priceValidUntil = product_data.get('offers', {}).get('priceValidUntil', '')
                    available = product_data.get('offers', {}).get('availability', '').replace('http://schema.org/', '')
                    variant_url = product_data.get('offers', {}).get('url', '')
                    rating = product_data.get('aggregateRating', {}).get('ratingValue', '')

                    try:
                        select_size = \
                            tree.xpath("//span[contains(@class, 'pro-select-size-title')]/following-sibling::span")[
                                0].text.strip()
                    except IndexError:
                        select_size = ''
                    try:
                        modelNumber = \
                            tree.xpath("//span[@class='mb-0 d-block pro-name']")[0].text.strip()
                    except IndexError:
                        modelNumber = ''

                    unprocessed_json = {"priceValidUntil": priceValidUntil,
                                        "availablity": available.replace('https://schema.org/', ''),
                                        "variant_url": variant_url, "rating": rating}

                    other = tree.xpath("//div[@class='pro-technical-specification-item mb-3']")

                    for i in range(len(other)):
                        key = tree.xpath("//div[@class='pro-technical-specification-item mb-3']")[i].text.strip().replace(':', '')
                        value_el = tree.xpath(f"(//div[@class='pro-technical-specification-item mb-3'])[{i + 1}]/a")
                        if len(value_el) > 0:
                            value = value_el[0].text.strip()
                            unprocessed_json[key] = value
                        else:
                            value = \
                                tree.xpath(f"(//div[@class='pro-technical-specification-item mb-3'])[{i + 1}]/span")[0].text.strip()
                            unprocessed_json[key] = value


                    parsed_data = {
                        'time_stamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        'tenant_id': "Cool Frames",
                        'competitor_name': "SmartBuyGlasses",
                        'sku': sku,
                        'title': name,
                        'competitor_url': site_url,
                        'brand': brand,
                        'category': category,
                        'sub_category': sub_category,
                        'modelNumber': modelNumber,
                        'description': description,
                        'size': size,
                        'regular_price': regular_price,
                        'discounted_price': discount_price,
                        "image_url": image_urls,
                        "priceCurrency": priceCurrency,
                    }
                    frame_color = unprocessed_json.get("Frame Color", None)
                    parsed_data['select_size'] = select_size
                    parsed_data["Frame Color"] = frame_color
                    if "Frame Color" in unprocessed_json:
                        del unprocessed_json["Frame Color"]
                    parsed_data["unprocessed_json"] = unprocessed_json

                    append_dict_to_list_and_save(parsed_data, "smartbuyglasses_data.json")
                    print(parsed_data)



urls = ["https://www.smartbuyglasses.com/designer-eyeglasses/SmartBuy-Collection/SmartBuy-Collection-Bibeth-T3032-C5-551748.html",
"https://www.smartbuyglasses.com/designer-eyeglasses/Tom-Ford/Tom-Ford-FT5634-B-Blue-Light-Block-001-420127.html",
"https://www.smartbuyglasses.com/designer-eyeglasses/SmartBuy-Collection/SmartBuy-Collection-Change-2421-C5-521084.html"
"https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/1-Day-Acuvue-Moist-30-Pack/442.html",
"https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/1-Day-Acuvue-Moist-90-Pack/448.html",
        "https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/1-Day-Acuvue-Moist-Multifocal-90-Pack/534.html",
        "https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/1-Day-Acuvue-Moist-for-Astigmatism-30-Pack/447.html",
        "https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/Acuvue-2-6-Pack/446.html",
        "https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/Acuvue-Oasys-1-Day-90-Pack/531.html",
        "https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/Acuvue-Oasys-1-Day-for-Astigmatism-30-Pack/706.html",
        "https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/Acuvue-Oasys-12-Pack/490.html",
        "https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/Acuvue-Oasys-Max-1-Day-30-Pack/850.html",
        "https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/Acuvue-Oasys-Max-1-Day-Multifocal-30-Pack/851.html",
        "https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/Acuvue-Vita-12-Pack/861.html",
        "https://www.smartbuyglasses.com/contact-lenses/daily-disposable-lenses/Acuvue-Vita-6-Pack/862.html"]

for row_num in range(1,max_row+1):
    url = sheet.cell(row=row_num, column=1).value  # Column A is index 1
    data_parser(url)