import openpyxl
import time
import os
from selenium import webdriver
from selenium.common import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from unidecode import unidecode
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
chrome_options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=chrome_options)
wait = WebDriverWait(driver, 10)

def ScrapData(i, url,count,count_1 , sequence, path, requirementList):
    requirementList = requirementList
    path = path
    # Comon data keep here
    allWidth =wait.until(expected_conditions.presence_of_element_located((By.XPATH, "//select[@id='widthWholeDropDown']")))
    allheight = wait.until(expected_conditions.presence_of_element_located((By.XPATH, "//select[@id='wholeHeightDropdown']")))
    allWidthText = allWidth.text.split('\n')
    allheightText = allheight.text.split('\n')
    allWidthSelect = Select(allWidth)
    allheightSelect = Select(allheight)
    # requirementList = ["24 x 36", "36 x 54", "36 x 60", "36 x 72", "48 x 48", "78 x 84"]
    for req in requirementList:
        if req[0:2] in allWidthText and req[5:7] in allheightText:
            # Handel stale element reference exception.
            try:
                allWidthSelect.select_by_visible_text(req[0:2])
            except StaleElementReferenceException:
                allWidth = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.XPATH, "//select[@id='widthWholeDropDown']")))
                allWidthText = allWidth.text.split('\n')
                allWidthSelect = Select(allWidth)
                allWidthSelect.select_by_visible_text(req[0:2])
            try:
                allheightSelect.select_by_visible_text(req[5:7])
            except StaleElementReferenceException:
                allheight = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.XPATH, "//select[@id='wholeHeightDropdown']")))
                allheightText = allheight.text.split('\n')
                allheightSelect = Select(allheight)
                allheightSelect.select_by_visible_text(req[5:7])
            time.sleep(2)
            driver.refresh()
            # If there are //a[text() = 'More Rows of Colors'] elements then click it.
            try:
                while True:
                    wait.until(EC.visibility_of_element_located((By.XPATH, "//a[text() = 'More Rows of Colors']"))).click()
            except:
                pass

            # Here below code will identify that how many color and texture realated comobination can be made.
            totalCombination = driver.find_elements(By.XPATH, "//div[@id='gcc-pip-swatches']/ul/li")
            if len(totalCombination) == 1:
                totalColor = driver.find_elements(By.XPATH, "//figure[@class='relative']")
                for clr in totalColor:
                    try:
                        clr.click()
                    except:
                        pass
                    count += 1

                    colorName = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'Selected Color')]/following-sibling::div"))).text.strip()
                    html = driver.page_source
                    with open(f'{path}{sequence}_{req[0:2]}_{req[5:7]}_{colorName.replace(" ", "_").replace("/", "").replace('"', '')}.html', 'w', encoding='utf-8') as f:
                        f.write(unidecode(html))
                    print("Dumps Saved for Url : ", url)
            else:
                print("Else Block Executed")
                colorList = driver.find_elements(By.XPATH, "//div[@id='gcc-pip-swatches']/ul/li[1]//ul/li/div//figcaption")
                textutrList = driver.find_elements(By.XPATH, "//div[@id='gcc-pip-swatches']/ul/li[2]//ul/li/div//figcaption")

                for clr in colorList:
                    try:
                        clr.click()
                    except:
                        pass
                    count += 1
                    for txt in textutrList:
                        txt.click()
                        count_1+=1
                        colorName = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'Selected Color')]/following-sibling::div"))).text.strip()
                        html = driver.page_source
                        textureName = txt.text.replace(" ", "-").replace("/", "").replace('"', '')
                        with open(f'{path}{sequence}_{req[0:2]}_{req[5:7]}_{colorName.replace(" ", "-").replace("/", "").replace('"', '')}_{textureName.replace("/", "").replace('"', '').replace(" ", "")}.html', 'w', encoding='utf-8') as f:
                            f.write(unidecode(html))
                        print("Dumps Saved for Url : ", url)
# Below function is used for different type of web page.

def scrapPattern(i, url,count,count_1, sequence, path, requirementList):
    requirementList = requirementList
    path = path
    time.sleep(1)
    print("Second Function Executed")
    widthInchDropdown = wait.until(EC.presence_of_element_located((By.XPATH, "//select[@aria-label='eighths']")))
    heightOptions = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Height')]/../..")))
    widthInchDropdownText = widthInchDropdown.text.split('\n')
    heightOptionsList = heightOptions.text.split('\n')[1:]
    widthInchSelect = Select(widthInchDropdown)
    # requirementList = ["24 x 36", "36 x 54", "36 x 60", "36 x 72", "48 x 48", "78 x 84"]
    for req in requirementList:
        if req[0:2] in widthInchDropdownText and req[5:7] in heightOptionsList:
            print(f"Combination Matched {req[0:2]} {req[5:7]}")
            # to handel stale elements for size selection
            try:
                widthInchSelect.select_by_visible_text(req[0:2])
                print(f"width selected {req[0:2]}")
            except StaleElementReferenceException:
                widthInchDropdown = driver.find_element(By.XPATH, "//select[@aria-label='eighths']")
                widthInchDropdownText = widthInchDropdown.text.split('\n')
                widthInchSelect = Select(widthInchDropdown)
                widthInchSelect.select_by_visible_text(req[0:2])
                print(f"Stale Element error has occoured but in except block width selected {req[0:2]}")
            # Below code will select height in case of box is there.
            driver.find_element(By.XPATH, f"//span[contains(text(),'Height')]/../../div[contains(text(),'{req[5:7]}')]").click()
            print(f"height clicked {req[5:7]}")
            time.sleep(2)
            driver.refresh()
            clrXpath = "//div[contains(@id,'Choice')]"
            colors = wait.until(EC.presence_of_all_elements_located((By.XPATH, clrXpath)))
            print(f"colors found {len(colors)}")
            for clr in colors:
                try:
                    clr.click()
                    print("color clicked")
                except:
                    pass
                    print("color not clicked not required already selected")
                count += 1
                time.sleep(5)
                html = driver.page_source
                colorName = wait.until(EC.presence_of_element_located((By.XPATH, f"//h2[text()='Choose your color' or contains(text(), 'Color')]/div/span"))).text.strip()
                with open(
                        f'{path}{sequence}_{req[0:2]}_{req[5:7]}_{colorName.replace(" ", "-").replace("/", "").replace('"', '')}.html', 'w', encoding='utf-8') as f:
                    f.write(unidecode(html))



# Open Xls File which contains urls.
workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet']
max_row = sheet.max_row
# Prepare a list of urls and sequence. which we were doing copy paste previously
url_list = []
sequence_list = []
for row_num in range(1, max_row + 1):
    url = sheet.cell(row=row_num, column=2).value
    sequence = sheet.cell(row=row_num, column=1).value
    url_list.append(url)
    sequence_list.append(sequence)
# Function to create name for the perticular url with the name of row number of url in excel sheet
def create_folder_if_not_exists(folder_path):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"Folder created at: {folder_path}")
    else:
        print(f"Folder already exists at: {folder_path}")

for i in range(0, max_row+1):
    if i in [88]:
        count = 0
        count_1 = 0
        url = url_list[i-1]
        sequence = sequence_list[i-1]
        path = f'D:\\dumps_blinds\\{i}/'
        create_folder_if_not_exists(path)
        requirementList = ["24 x 36", "36 x 54", "36 x 60", "36 x 72", "48 x 48", "78 x 84"]
        try:
            driver.get(url)
            time.sleep(5)
            if "?sku=" not in url:
                ScrapData(i, url,count,count_1, sequence, path, requirementList)
                with open(f'successURLLatest1.txt', 'a') as f1:
                    f1.write(url)
                    f1.write('\n')
            else:
                scrapPattern(i, url,count,count_1, sequence, path, requirementList)
                with open(f'successURLLatest1.txt', 'a') as f2:
                    f2.write(url)
                    f2.write('\n')
        except Exception as e:
            print(e)
            with open(f'errorURLLatest1.txt', 'a') as f3:
                f3.write(url)
                f3.write('\n')

driver.quit()