import json
import os

import openpyxl
from scrapy import Selector
from datetime import datetime
path = "E:\\dumps_tractor"

path_list = []
for file in os.listdir(path):
    if file.endswith(".html"):
        path_list.append(os.path.join(path, file))

count = 1

workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

for i in path_list[:]:
    # print(count, end=" ")
    with open(i, 'r', encoding='utf-8') as f:
        html = f.read()
    selector = Selector(text=html)
    name = selector.xpath("//meta[@itemprop='name']/@content").get()
    url = selector.xpath("//meta[@property='og:url']/@content").get()
    sku = selector.xpath("//span[@id='sku']/text()").get()

    Level1 = None
    Level2 = None
    Level3 = None
    Level4 = None
    Level5 = None

    breadcrumb = selector.xpath("//nav[@id='breadcrumb_generic']//li//span/text()").getall()
    # if len(breadcrumb) == 5:
    #     print(breadcrumb)
    if len(breadcrumb) > 0:
        if len(breadcrumb) == 4:
            Level1 = breadcrumb[1]
            Level2 = breadcrumb[2]
        elif len(breadcrumb) == 5:
            Level1 = breadcrumb[1]
            Level2 = breadcrumb[2]
            Level3 = breadcrumb[3]
        elif len(breadcrumb) == 6:
            Level1 = breadcrumb[1]
            Level2 = breadcrumb[2]
            Level3 = breadcrumb[3]
            Level4 = breadcrumb[4]
        elif len(breadcrumb) == 7:
            Level1 = breadcrumb[1]
            Level2 = breadcrumb[2]
            Level3 = breadcrumb[3]
            Level4 = breadcrumb[4]
            Level5 = breadcrumb[5]
    description = selector.xpath("//meta[@name='description']/@content").get()
    regular_price = None
    discounted_price = None
    price1 = selector.xpath("//div[contains(@class, 'pricing-content')]//span[@class='offer_price_min mvPrice']/text()").get()
    price2 = selector.xpath("//div[@class='pricing-content']//span[@class='list_price_min']/text()").get()

    if price1 is not None and price2 is None:
        regular_price = price1.replace("$", "")
        discounted_price = None
    elif price1 is not None and price2 is not None:
        regular_price = price2.replace("$", "")
        discounted_price = price1.replace("$", "")
    elif price1 is None and price2 is not None:
        regular_price = price2.replace("$", "")
        discounted_price = None

    unprocessed_json = {}
    images = []
    image_list = selector.xpath("//div[contains(@class, 'image-item slick-slide')]/img")
    for im in image_list:
        imageurl = im.xpath("./@data-lazy").get()
        if imageurl is None:
            imageurl = im.xpath("./@src").get()
        images.append(imageurl)

    elements = selector.xpath("//div[@id ='specifications']//tr")
    for el in elements[1:]:
        key = el.xpath("./td[1]/text()").get().replace("\n\t\t\t\t\t\t\t\t\t\t\t", " ")
        value = el.xpath("./td[2]/text()").get()
        unprocessed_json[key] = value
    rating = selector.xpath("(//span[contains(@title, 'Product Rating')])[2]/text()").get()

    upc = sheet.cell(row=int(i.replace("E:\\dumps_tractor\\", "").replace(".html", "")), column=1).value

    upc = upc.replace("https://www.tractorsupply.com/tsc/search/", "")

    print(i)

    parsed_data = {'COMP_ITEM_DESCRIPTION': name,
                   'COMP_SKU': sku,
                   'COMP_REGULAR_PRICE': regular_price,
                   'COMP_PROMO_PRICE': discounted_price,
                   'COMP_UDA_LEVEL1': Level1,
                   'COMP_UDA_LEVEL2': Level2,
                   'COMP_UDA_LEVEL3': Level3,
                   'COMP_UDA_LEVEL4': Level4,
                   'COMP_UDA_LEVEL5': Level5,
                   "COMP_UPC": upc,
                   "COMP_ITEM_URL": url,
                   "COMP_CURRENCY": "USD",
                   'COMP_UDA_ATTRIBUTE2': rating,
                   "COMP_ITEM_LONG_DESCRIPTION": description,
                   "COMP_ITEM_IMAGE_URL": images,
                   "COMP_UDA_ATTRIBUTE4": {**unprocessed_json}}

    # print(parsed_data)
    json_file_path = 'Tractor.json'
    # Function to append a dictionary to a list and save it to a JSON file
    def append_dict_to_list_and_save(dictionary, file_path):
        try:
            # Read the existing data from the JSON file (if it exists)
            with open(file_path, 'r') as file:
                data = json.load(file)
        except (FileNotFoundError, json.JSONDecodeError):
            # If the file doesn't exist or is empty, create an empty list
            data = []
        # Append the new dictionary to the list
        data.append(dictionary)
        # Save the updated data back to the JSON file
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=4)
    # Dictionary to append to the list
    append_dict_to_list_and_save(parsed_data, json_file_path)
    count += 1