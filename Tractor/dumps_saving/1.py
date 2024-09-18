import time


from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
from seleniumbase import Driver
from unidecode import unidecode
driver = Driver(browser='Chrome')
wait = WebDriverWait(driver, 1)


def get_data():
    html_path = 'E:\\dumps_tractor/'
    html = driver.page_source
    with open(f'{html_path}{row_num}.html', 'w', encoding='utf-8') as f:
        f.write(unidecode(html))


workbook = openpyxl.load_workbook("../urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row


for row_num in range(1, max_row+1):
    url = sheet.cell(row=row_num, column=2).value
    driver.get(url)
    time.sleep(5)

    try:
        get_data()
    except:
        time.sleep(10)
        driver.get(url)
        get_data()
    print(row_num, url)