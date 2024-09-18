import json
from datetime import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By

driver = webdriver.Edge()
wait = WebDriverWait(driver, 3)

#Open Xls File which contains urls.
workbook = openpyxl.load_workbook("Healthforyou.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row
# Read data from column A (column index 1) line by line
for row_num in range(1, max_row+1):
    url = sheet.cell(row=row_num, column=1).value  # Column A is index 1
    driver.get(url)
    script_tag  = driver.find_element(By.XPATH, "//script[@type='application/ld+json' and contains(.,'sku')]")
    json_object = script_tag.get_attribute('innerHTML')
    product_data = json.loads(json_object, strict=False)
    category = driver.find_element(By.XPATH, "(//li[@class='breadcrumb-item'])[1]").text
    sub_category = driver.find_element(By.XPATH, "(//li[@class='breadcrumb-item'])[2]").text
    name = product_data.get('name')
    brand = product_data.get('brand').get("name")
    description = product_data.get('description')
    url = product_data.get('url', '')

    main_sku = product_data.get('sku', '')
    if "aggregateRating" in product_data:
        rating = product_data.get('aggregateRating', {}).get('ratingValue', '')
        rated_by = product_data.get('aggregateRating', {}).get('reviewCount', '')
    else:
        rating = ""
        rated_by = ""
    variant = product_data.get('offers')
    if type(variant) == list:
        for i in range(len(variant)):
            timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            priceCurrency = variant[i].get('priceCurrency', '')
            price = variant[i].get('price', '')
            priceValidUntil = variant[i].get('priceValidUntil', '')
            mpn_sku = variant[i].get('itemOffered',{}).get("mpn", '')
            variant_sku = variant[i].get('itemOffered', {}).get("sku", '')
            itemCondition = variant[i].get('itemCondition', '')
            availability = variant[i].get('availability','')
            size = driver.find_element(By.XPATH, "(//div[contains(@class, 'siz-chart')])["+str(i+1)+"]/div/h4").text

            pack_info = variant[i].get('itemOffered', {}).get("name", '').split(",")[-1].replace(" ","")

            parsed_data = {'time_stamp': timestamp_str, 'tenant_id': 'VITALITY_MEDICAL', 'competitor_id': '', 'competitor_name': 'HealthProductsForYou', 'upc_code': '', 'sku': mpn_sku, 'title': name, 'url': "https://www.healthproductsforyou.com"+url, 'brand': brand, 'category': category, 'description': description, 'quantity': '', 'size': size, 'weight': '', 'regular_price': price, 'discount': '', 'unprocessed_json': {"pack_info":pack_info,"rating_value":rating, "rated_by":rated_by, 'sub_category':sub_category, "priceCurrency":priceCurrency, "priceValidUntil":priceValidUntil, "variant_sku":variant_sku, "itemCondition":itemCondition.replace("http://schema.org/", ""), "availability":availability.replace("http://schema.org/", "").replace("https://schema.org/", "")},}

            json_file_path = 'Healthdata.json'
            # Function to append a dictionary to a list and save it to a JSON file
            def append_dict_to_list_and_save(dictionary, file_path):
                try:
                    # Read the existing data from the JSON file (if it exists)
                    with open(file_path, 'r') as file:
                        data = json.load(file)
                except (FileNotFoundError, json.JSONDecodeError):
                    # If the file doesn't exist or is empty, create an empty list
                    data = []
                # Append the new dictionary to the list
                data.append(dictionary)
                # Save the updated data back to the JSON file
                with open(file_path, 'w') as file:
                    json.dump(data, file, indent=4)
            # Dictionary to append to the list
            append_dict_to_list_and_save(parsed_data, json_file_path)
            print(parsed_data)
    else:
        timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        priceCurrency = variant.get('priceCurrency', '')
        price = variant.get('price', '')
        priceValidUntil = variant.get('priceValidUntil', '')
        mpn_sku = variant.get('itemOffered',{}).get("mpn", '')
        variant_sku = variant.get('itemOffered', {}).get("sku", '')
        itemCondition = variant.get('itemCondition', '')
        availability = variant.get('availability','')
        pack_info = ""
        try:
            size = driver.find_element(By.XPATH, "(//div[contains(@class, 'siz-chart')])["+str(1)+"]/div/h4").text
        except:
            size = ""
        parsed_data = {'time_stamp': timestamp_str, 'tenant_id': 'VITALITY_MEDICAL', 'competitor_id': '', 'competitor_name': 'HealthProductsForYou', 'upc_code': '', 'sku': mpn_sku, 'title': name, 'url': "https://www.healthproductsforyou.com"+url, 'brand': brand, 'category': category, 'description': description, 'quantity': '', 'size': size, 'weight': '', 'regular_price': price, 'discount': '', 'unprocessed_json': {"pack_info":pack_info, "rating_value":rating, "rated_by":rated_by, "priceCurrency":priceCurrency, "priceValidUntil":priceValidUntil, "variant_sku":variant_sku, "itemCondition":itemCondition.replace("http://schema.org/", ""), "availability":availability.replace("http://schema.org/", "")},};
        json_file_path = 'Healthdata.json'
        # Function to append a dictionary to a list and save it to a JSON file
        def append_dict_to_list_and_save(dictionary, file_path):
            try:
                # Read the existing data from the JSON file (if it exists)
                with open(file_path, 'r') as file:
                    data = json.load(file)
            except (FileNotFoundError, json.JSONDecodeError):
                # If the file doesn't exist or is empty, create an empty list
                data = []
            # Append the new dictionary to the list
            data.append(dictionary)
            # Save the updated data back to the JSON file
            with open(file_path, 'w') as file:
                json.dump(data, file, indent=4)
        # Dictionary to append to the list
        append_dict_to_list_and_save(parsed_data, json_file_path)
        print(parsed_data)
