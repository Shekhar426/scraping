import json
from datetime import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

chrome_options = Options()
chrome_options.add_argument("--headless")

driver = webdriver.Chrome()
wait = WebDriverWait(driver, 10)

#Open Xls File which contains urls.
workbook = openpyxl.load_workbook("Urls.xlsx")
sheet = workbook['Sheet']
max_row = sheet.max_row

for row_num in range(1,50):
    url = sheet.cell(row=row_num, column=1).value  # Column A is index
    driver.get(url)
    title = driver.title
    if title.endswith("Not Found"):
        print(url, "Product Not Found")
    else:
        script_tag  = driver.find_element(By.XPATH, "//script[@type='application/ld+json' and contains(.,'sku')]")
        json_object = script_tag.get_attribute('innerHTML')
        product_data = json.loads(json_object, strict=False)
        item_name = product_data.get('name')
        sku = product_data.get('sku')
        url =  product_data.get('url')
        brand = product_data.get('brand').get("name")
        category = driver.find_element(By.XPATH, "(//a[@class='breadcrumb-label'])[2]").text.replace(".", "")
        price = product_data.get('offers').get('price')
        price_currency = product_data.get('offers').get('priceCurrency')
        availablity = product_data.get('offers').get('availability').replace("https://schema.org/", "")
        price_valid_until = product_data.get('offers').get('priceValidUntil')

        # Will get data from dynamic list
        details = driver.find_elements(By.XPATH, "//dl[@class='productView-info']/div/dt")
        details_list = {}
        for i in range(len(details)):
            key = driver.find_element(By.XPATH, f"//dl[@class='productView-info']/div[{i+1}]/dt").text.replace(":", "").replace(".", "")
            value = driver.find_element(By.XPATH, f"//dl[@class='productView-info']/div[{i+1}]/dd").text.replace(".", "")
            details_list[key] = value
        upc = details_list.get("UPC")
        # Get data from static list
        raw_text = driver.find_element(By.XPATH, "//div[@class='tabs-contents']").text.split("\n")
        for i in range(len(raw_text)):
            if ":" in raw_text[i] and "Features and Specifications" not in raw_text[i] and "NOTE:" not in raw_text[i]:
                key_value = raw_text[i].split(":")
                details_list[key_value[0].replace(".", "")] = key_value[1].replace(".", "")
        parsed_data = {'time_stamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'tenant_id': 'Primary Arms', 'competitor_id': '', 'competitor_name': 'Cheaper than dirt', 'upc_code': upc, 'sku': sku, 'title': item_name, 'url': driver.current_url, 'brand': brand, 'category': category, 'description': '', 'quantity': '', 'size': '', 'weight': "", 'regular_price': price, 'discount': '', 'unprocessed_json': {"price_currency":price_currency, "price_valid_until":price_valid_until, "availablity":availablity, **details_list}}
        print(parsed_data)
        json_file_path = 'cheaperthandirt.json'
        # Function to append a dict to a list and save it to a JSON file
        def append_dict_to_list_and_save(dictionary, file_path):
            try:
                # Read the existing data from the JSON file (if it exists)
                with open(file_path, 'r') as file:
                    data = json.load(file)
            except (FileNotFoundError, json.JSONDecodeError):
                # If the file doesn't exist or is empty, create an empty list
                data = []
            data.append(dictionary)
            # Save the updated data back to the JSON file
            with open(file_path, 'w') as file:
                json.dump(data, file, indent=4)
        # Dictionary to append to the list
        append_dict_to_list_and_save(parsed_data, json_file_path)