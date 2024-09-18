import openpyxl
import os
import argparse
import random

from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from seleniumbase import Driver
from unidecode import unidecode
from completed import completed
from failure import failure
import time
import json

# import completed

# Argument parser to take range values from command line
parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument('--start', type=int, required=True, help='Starting row number')
parser.add_argument('--end', type=int, required=True, help='Ending row number')
args = parser.parse_args()

# Open Xls File which contains URLs.
workbook = openpyxl.load_workbook("/usr/src/app/urlsRH.xlsx")
sheet = workbook['Sheet1']

def create_folder_if_not_exists(folder_path):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"Folder created at: {folder_path}", end=" | ")
    else:
        print(f"Folder already exists at: {folder_path}")

def dumps_downloading(i):
    count = 0
    driver.get(url)
    wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='MuiGrid-root MuiGrid-container']//h1")))
    print(driver.title)
    Flag = False
    if "Access" in driver.title:
        Flag = True
    tryCount = 0
    while Flag == True:
        time.sleep(random.choice([2, 2, 4, 2, 3, 2, 2, 4, 2, 3, 2, 2, 4, 2, 3]))
        driver.reload()
        tryCount = tryCount + 1
        if "Access" not in driver.title:
            print("Access Denied Error is resolved  {}".format(i))
            Flag = False
            break
        elif tryCount > 10:
            print("Max Try for Access Denied Moving to next url  {}".format(i))
            Flag = False
            break
        print("Tried Count  {}. Please wait let's try few more time........".format(tryCount))
    wait.until(EC.visibility_of_element_located((By.XPATH, "//h1[contains(@class, 'MuiTypography-h1')]")))

    # Check that page is related to first condition. if page condition is of first type then set variable value as First
    try:
        arrowPointer = wait.until(EC.visibility_of_all_elements_located((By.XPATH,
                                                                         "//div[contains(@class, 'flex items-center justify-between pr-2 relative') and contains(@data-testid, 'SELECT')]")))
    except:
        arrowPointer = []
    try:
        typeOfGrid = wait.until(
            EC.visibility_of_all_elements_located((By.XPATH, "//div[@id='component-swatch-groups']")))
    except:
        typeOfGrid = []

    # If there are only colors and no arrow pointer available to click
    complete_data = []
    name = wait.until(
        EC.visibility_of_all_elements_located((By.XPATH, '//div[@class="MuiGrid-root MuiGrid-container"]/div/h1')))
    title_name = name[0].text.title()
    dumpCountFor_PerticularURL = 0
    print(title_name)
    if len(arrowPointer) == 0 and len(typeOfGrid) == 1:
        print("************ only finish")
        combinationElement = wait.until(
            EC.visibility_of_all_elements_located((By.XPATH, "//div[@id='component-swatch-groups']//ul/li")))
        # import pdb; pdb.set_trace()
        optionsName = wait.until(EC.visibility_of_all_elements_located((By.XPATH,
                                                                        f"//p[text()='{title_name}']/../../a/../../../../../following-sibling::div[1]//select[contains(@id,'Finish')]/option")))
        # print("Total options found", len(combinationElement), len(optionsName))

        if optionsName:
            dumpCountFor_PerticularURL = 0
            # print("Total combination found", len(combinationElement))
            for index in range(len(optionsName)):
                if optionsName[index].get_attribute('value') != '':
                    data = {
                        'url': url,
                        'Finish': optionsName[index].get_attribute('value'),
                        'Fabricorleather': None
                    }
                    print(data)
                    complete_data.append(data)
                dumpCountFor_PerticularURL += 1
            count += 1
            with open(f'/usr/src/app/json_rh/rhSuccess.txt', 'a') as f:
                f.write(url)
                f.write('\n')
        else:
            with open(f'/usr/src/app/json_rh/rhFailure.txt', 'a') as f3:
                f3.write(url)
                f3.write('\n')
            count += 1
            assert 1 == 2
    elif len(arrowPointer) == 1 and len(typeOfGrid) == 1:
        print("********************* Finish and Leatherfabric")
        # combinationElement = wait.until(EC.visibility_of_all_elements_located((By.XPATH, '//select[contains(@id,"Finish")]/option')))
        optionsName = wait.until(EC.visibility_of_all_elements_located((By.XPATH,
                                                                        f'//p[text()="{title_name}"]/../../a/../../../../../following-sibling::div[1]//select[contains(@id,"Finish")]/option')))
        dumpCountFor_PerticularURL = 0

        for index in range(len(optionsName)):
            if optionsName[index].get_attribute('value') != '':
                finish = optionsName[index].get_attribute('value')
                optionsName[index].click()
                checkleatherfabricLength = wait.until(EC.visibility_of_all_elements_located((By.XPATH,
                                                                                             f'//p[text()="{title_name}"]/../../a/../../../../../following-sibling::div[1]//select[contains(@id,"Fabric") or contains(@id,"Leather")]/option')))
                loopCount = len(checkleatherfabricLength)
                for j in range(loopCount):
                    if checkleatherfabricLength[j].get_attribute('value') != '':
                        leatherfabricName = checkleatherfabricLength[j].get_attribute('value')
                        data = {
                            'url': url,
                            'Finish': finish,
                            'Fabricorleather': leatherfabricName
                        }
                        print(data)
                        complete_data.append(data)
                        dumpCountFor_PerticularURL += 1
                        count += 1
                        with open(f'/usr/src/app/json_rh/rhSuccess.txt', 'a') as f:
                            f.write(url)
                            f.write('\n')
        count += 1

    elif len(arrowPointer) == 1 and len(
            typeOfGrid) == 0:  # Third condition where there is one arrow pointer but no grids found
        print("************************ Only leatherfabric")
        checkleatherfabricLength = wait.until(EC.visibility_of_all_elements_located((By.XPATH,
                                                                                     f'//p[text()="{title_name}"]/../../a/../../../../../following-sibling::div[1]//select[contains(@id,"Fabric") or contains(@id,"Leather")]/option')))
        loopCount = len(checkleatherfabricLength)
        for j in range(loopCount):
            if checkleatherfabricLength[j].get_attribute('value') != '':
                leatherfabricName = checkleatherfabricLength[j].get_attribute('value')
                data = {
                    'url': url,
                    'Finish': None,
                    'Fabricorleather': leatherfabricName
                }
                print(data)
                complete_data.append(data)
                dumpCountFor_PerticularURL += 1
        count += 1
        with open(f'/usr/src/app/json_rh/rhSuccess.txt', 'a') as f:
            f.write(url)
            f.write('\n')
        count += 1

    elif len(arrowPointer) == 0 and len(typeOfGrid) == 0:
        print("******************** No options")
        data = {
            'url': url,
            'Finish': None,
            'Fabricorleather': None
        }
        print(data)
        complete_data.append(data)

        count += 1
        with open(f'/usr/src/app/json_rh/rhSuccess.txt', 'a') as f:
            f.write(url)
            f.write('\n')

    output_file = f"/usr/src/app/json_rh/rh_variants_{args.start}.json"
    with open(output_file, 'a') as file:
        json.dump(complete_data, file, indent=4)

# Initialize Driver with headless option
driver = Driver(uc=True, headed=False, browser='chrome', headless=True)
wait = WebDriverWait(driver, 20)

for i in range(args.start, args.end + 1):
    # if i not in completed.completed:
    sequence = sheet.cell(row=i, column=1).value
    url = sheet.cell(row=i, column=1).value

    if url not in completed:
        print("Row Number : {} {}".format(i, url), end=" | ")
        try:
            dumps_downloading(i)
        except Exception as e:
            print(e)
            with open(f'/usr/src/app/json_rh/rhFailure.txt', 'a') as f3:
                f3.write(url)
                f3.write('\n')
            print(f"Error in URL Apppended in Logs {url}")
    elif url in completed:
        print(f"Url is already Completed {url}")
    elif url in failure:
        print(f"Url is already Failed {url}")