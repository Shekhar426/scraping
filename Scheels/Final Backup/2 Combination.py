import requests
from bs4 import BeautifulSoup
import openpyxl

# Function to get data using requests and BeautifulSoup
def get_data(site_url, row_num):
    try:
        response = requests.get(site_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            first_ul = soup.find('ul', {'role': 'presentation'})
            if first_ul:
                total_variation_type = first_ul.find_all('li', recursive=False)
            else:
                total_variation_type = []
            if len(total_variation_type) == 2:
                pid = soup.select_one("span.productsku").text
                print("Entered in ELIF 2")
                try:
                    optclr_elements = soup.select("ul[role='presentation'] li:nth-of-type(1) .value a")
                    if len(optclr_elements) != 2:
                        raise ValueError(f"Expected 2 elements, but found {len(optclr_elements)} for {site_url}")
                    optclr1 = soup.select("ul[role='presentation'] li:nth-of-type(1) .value a")
                    optsize2 = soup.select("ul[role='presentation'] li:nth-of-type(2) .value a")
                    base_url = "https://www.scheels.com/on/demandware.store/Sites-scheels-Site/en_US/Product-Variation?"

                    product_urlsoptclr_List = []
                    product_urlsoptsize_List = []

                    # Parse options for the first and second variations
                    for optclr in optclr1:
                        optclr_params = optclr.get('data-href').split('?')[1].split('&')
                        for puoc in optclr_params:
                            if "pid=" not in puoc and puoc not in product_urlsoptclr_List:
                                product_urlsoptclr_List.append(puoc)
                    for optsize in optsize2:
                        optsize_params = optsize.get('data-href').split('?')[1].split('&')
                        for puos in optsize_params:
                            if "pid=" not in puos and puos not in product_urlsoptsize_List:
                                product_urlsoptsize_List.append(puos)
                    for clrs in product_urlsoptclr_List:
                        for sizes in product_urlsoptsize_List:
                            customized_url = f"{base_url}pid={pid}&{clrs}&{sizes}"
                            with open('productUrls.txt', 'a') as f5:
                                f5.write(customized_url + '\n')
                except:
                    pass
        elif response.status_code != 200:
            with open('responseError.txt', 'a') as f:
                f.write(site_url + '\n')
    except:
        with open('Error.txt', 'a') as f:
            f.write(site_url + '\n')
workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

for row_num in range(2, 3):
    url = sheet.cell(row=row_num, column=1).value
    get_data(url, row_num)
    print(f"{row_num} {url}")
