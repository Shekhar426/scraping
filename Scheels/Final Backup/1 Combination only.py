import requests
from bs4 import BeautifulSoup
import openpyxl

# Function to get data using requests and BeautifulSoup
def get_data(site_url, row_num):
    try:
        response = requests.get(site_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            first_ul = soup.find('ul', {'role': 'presentation'})
            if first_ul:
                total_variation_type = first_ul.find_all('li', recursive=False)
            else:
                total_variation_type = []
            # total_variation_type = soup.select("ul[role='presentation']:nth-of-type(1) > li")

            if len(total_variation_type) == 1:
                # Corrected the selector to access the first `ul` with `role='presentation'` and then select `li a` elements
                ul_elements = soup.select("ul[role='presentation']")
                if len(ul_elements) > 0:
                    Options1 = ul_elements[0].select("li a")
                    for opt1Urls in Options1:
                        productUrlsopt1 = opt1Urls.get('data-href')
                        if productUrlsopt1:
                            with open('productUrls.txt', 'a') as f2:
                                f2.write(productUrlsopt1 + '\n')
        elif response.status_code != 200:
            with open('responseError.txt', 'a') as f:
                f.write(site_url + '\n')
    except:
        with open('Error.txt', 'a') as f:
            f.write(site_url + '\n')
workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

for row_num in range(10, 20):
    url = sheet.cell(row=row_num, column=1).value
    get_data(url, row_num)
    print(f"{row_num} {url}")
