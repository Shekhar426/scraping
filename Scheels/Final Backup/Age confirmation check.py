import requests
from bs4 import BeautifulSoup
import openpyxl

# Function to get data using requests and BeautifulSoup
def get_data(site_url, row_num):
    try:
        response = requests.get(site_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            # Find the title tag
            title = soup.title.string if soup.title else ""
            # Check if the title text matches the specific string
            if title.strip() == "Age Confirmation | SCHEELS.com":
                # Write the URL to AgeBlocker.txt
                with open("AgeBlocker.txt", "a") as file:
                    file.write(url + "\n")
            else:
                # Write the URL to normal.txt
                with open("normal.txt", "a") as file:
                    file.write(url + "\n")
        else:
            with open('Not_200.txt', 'a') as f:
                f.write(site_url + '\n')
    except:
        with open('Error.txt', 'a') as f:
            f.write(site_url + '\n')
workbook = openpyxl.load_workbook("../Rough/urls.xlsx")
sheet = workbook['Sheet2']
max_row = sheet.max_row

for row_num in range(1, 3000):
    url = sheet.cell(row=row_num, column=1).value
    get_data(url, row_num)
    print(f"{row_num} {url}")
