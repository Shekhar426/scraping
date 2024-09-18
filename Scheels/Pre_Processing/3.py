from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl
from seleniumbase import Driver

driver = Driver(uc=True, headed=False, browser='chrome', headless=True)
wait = WebDriverWait(driver, 1)

def get_data(site_url, row_num):
    driver.get(site_url)
    try:
        # In this block check if product is not found
        driver.find_element(By.XPATH, "//li[contains(text(), 'The link may be incorrect or old')]")
        with open('NotFound.txt', 'a') as f:
            f.write(site_url + '\n')
        print(f"{row_num} {site_url} Product Not Found")
    except:
        pid = driver.execute_script("return document.querySelector(\"span.productsku\").textContent;")
        try:
            totalVariationType = driver.find_elements(By.XPATH, "(//ul[@role='presentation'])[1]/li")
            if len(totalVariationType) == 0:
                print("Entered in IF")
                productUrls = site_url
                with open('productUrls.txt', 'a') as f:
                    f.write(productUrls + '\n')
                with open('success105959.txt', 'a') as f3:
                    f3.write(site_url + '\n')
                print(f"{row_num} {site_url}")
            elif len(totalVariationType) == 1:
                print("Entered in ELIF 1")
                Options1 = driver.find_elements(By.XPATH, "(//ul[@role='presentation'])[1]/li//a")
                for opt1Urls in Options1:
                    productUrlsopt1 = (opt1Urls.get_attribute('data-href'))
                    with open('productUrls.txt', 'a') as f2:
                        f2.write(productUrlsopt1 + '\n')
                with open('success105959.txt', 'a') as f4:
                    f4.write(site_url + '\n')
                print(f"{row_num} {site_url} Variation : {len(totalVariationType)} Options : {len(Options1)}")

            elif len(totalVariationType) == 2:
                print("Entered in ELIF 2")
                try:
                    assert len(driver.find_elements(By.XPATH, "(//ul[@role='presentation'])[1]/li[1]//div[@class='value']")) == 2
                    optclr1 = driver.find_elements(By.XPATH, "(//ul[@role='presentation'])[1]/li[1]//div[@class='value']//a")
                    optsize2 = driver.find_elements(By.XPATH, "(//ul[@role='presentation'])[1]/li[2]//div[@class='value']//a")
                    baseUrl = "https://www.scheels.com/on/demandware.store/Sites-scheels-Site/en_US/Product-Variation?"
                    productUrlsoptclr_List = []
                    productUrlsoptsize_List = []
                    for optclr in optclr1:
                        for optsize in optsize2:
                            productUrlsoptclr = (optclr.get_attribute('data-href')).split('?')[1].split('&')
                            for puoc in productUrlsoptclr:
                                if "pid=" not in puoc and puoc not in productUrlsoptclr_List:
                                    productUrlsoptclr_List.append(puoc)
                            productUrlsoptsize = (optsize.get_attribute('data-href')).split('?')[1].split('&')
                            for puos in productUrlsoptsize:
                                if "pid=" not in puos and puos not in productUrlsoptsize_List:
                                    productUrlsoptsize_List.append(puos)
                    for clrs in productUrlsoptclr_List:
                        for sizes in productUrlsoptsize_List:
                            customizedUrl = f"{baseUrl}pid={pid}&{clrs}&{sizes}"
                            with open('productUrls.txt', 'a') as f5:
                                f5.write(customizedUrl + '\n')
                    with open('success105959.txt', 'a') as f6:
                        f6.write(site_url + '\n')
                    print(f"{row_num} {site_url} Variation : {len(totalVariationType)} Options : {len(productUrlsoptclr_List) * len(productUrlsoptsize_List)}")
                except:
                    print(f"{row_num} {site_url} 2 varient came but scenario was not handled Failed")
                    with open('Not Handled.txt', 'a') as f7:
                        f7.write(site_url + '\n')
            else:
                print("Entered in Else")
                print(f"{row_num} {site_url} Variation : Pending Options : Pending" )
                with open('Not Handled.txt', 'a') as f8:
                    f8.write(site_url + '\n')
        except:
            with open('Error.txt', 'a') as f9:
                f9.write(url + '\n')

workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

driver.get('https://www.scheels.com')
driver.find_element(By.XPATH, "//input[@id='us-site']").send_keys(Keys.ESCAPE)

count = 0 # Counting number of age confirmation popup

for row_num in range(11639, 15000):
    url = sheet.cell(row=row_num, column=1).value
    if count == 0:
        pass
    try:
        get_data(url, row_num)
    except:
        with open('Error.txt', 'a') as f:
            f.write(url + '\n')