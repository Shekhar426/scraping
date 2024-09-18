import openpyxl

workbook = openpyxl.load_workbook("../Rough/urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

List1 = []
List2 = []

for row_num in range(1, max_row+1):
    url = sheet.cell(row=row_num, column=1).value
    List1.append(url)

workbook2 = openpyxl.load_workbook("../Rough/urls.xlsx")
sheet2 = workbook2['Sheet2']
max_row2 = sheet2.max_row

for row_num in range(1, max_row2+1):
    url2 = sheet2.cell(row=row_num, column=1).value
    List2.append(url2)

for urls in List1:
    if urls not in List2:
        with open("urls.txt", "a") as f:
            f.write(f"{urls}\n")