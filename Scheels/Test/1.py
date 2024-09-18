import json

import requests
from scrapy import Selector
import openpyxl

# Function to get data using requests and BeautifulSoup
def get_data(site_url, row_num):
    try:
        response = requests.get(site_url)
        if response.status_code == 200:
            selector = Selector(text=response.text)
            unprocessed_json = {}
            name = selector.xpath("//h1[@class='product-name-main']/text()").get()
            if name is not None:
                name = name.replace("\n", "")
            brand = selector.xpath("//h1[@class='product-name-main']/../a/text()").get()
            url = selector.xpath("//link[@rel='canonical']/@href").get()
            sku = selector.xpath("//span[@class='productsku']/text()").get()
            description = selector.xpath("//meta[@name='description']/@content").get()

            category = None
            sub_category = None
            Levels = selector.xpath("//nav[@aria-label= 'breadcrumbs']//span")
            for level in range(len(Levels)):
                if level == 0:
                    category = Levels[level].xpath("text()").get()
                elif level == 1:
                    sub_category = Levels[level].xpath("text()").get()

            regular_price = None
            discounted_price = None
            price1 = selector.xpath("(//span[@class='price-sales'])[1]/span//span[last()]/text()").get()
            price2 = selector.xpath("(//del[@class='price-standard sale'])[1]/text()").get()
            if price1 is not None and price2 is not None:
                regular_price = price2
                discounted_price = price1
            if price1 is not None and price2 is None:
                regular_price = price1
                discounted_price = None
            if regular_price is not None:
                regular_price = regular_price.replace("\n", "").replace("$", "")
            if discounted_price is not None:
                discounted_price = discounted_price.replace("\n", "").replace("$", "")

            variations = selector.xpath("(//ul[@role='presentation'])[1]/li//div[@class='label']")

            for var in variations:
                key = var.xpath("./h2/span/text()").get()
                value = []
                values = var.xpath("./following-sibling::div[@class='value']//ul/li/a")
                for vl in values:
                    value.append(vl.xpath("./@data-label").get())
                unprocessed_json[key] = value

            currency = selector.xpath("//meta[@itemprop='priceCurrency']/@content").get()
            Availability = selector.xpath("//meta[@itemprop='availability']/@content").get()
            if Availability is not None:
                Availability = Availability.replace("http://schema.org/", "")

            images = []
            image_list = selector.xpath("(//div[contains(@class, 'product-image-container')]//ul)[1]//a")
            for img in image_list:
                imageurl = img.xpath("./@href").get()
                images.append(imageurl)
            images = list(set(images))

            elements = selector.xpath(
                "//div[@data-test-id='guided-accordion']//div[contains(@data-test-id, 'guided-accordion-item')]")
            for i in elements:
                key = i.xpath(".//div[@data-test-id='guided-accordion-header-title']/text()").get().strip().replace(
                    "\\n",
                    "")
                value = i.xpath(".//div[@data-test-id='guided-accordion-header-attribute-value']/text()").get()
                unprocessed_json[key] = value

            parsed_data = {
                'COMP_ITEM_DESCRIPTION': name,
                'COMP_BRAND': brand,
                'COMP_ITEM_LONG_DESCRIPTION': description,
                'COMP_SKU': sku,
                'COMP_UDA_LEVEL1': category,
                'COMP_UDA_LEVEL2': sub_category,
                'COMP_REGULAR_PRICE': regular_price,
                'COMP_PROMO_PRICE': discounted_price,
                "COMP_CURRENCY": currency,
                "AVAILABILITY": Availability,
                "COMP_ITEM_URL": url,
                "COMP_ITEM_IMAGE_URL": images,
                "COMP_UDA_ATTRIBUTE4": {**unprocessed_json}}
            json_file_path = 'Scheels.json'
            # Function to append a dictionary to a list and save it to a JSON file
            def append_dict_to_list_and_save(dictionary, file_path):
                try:
                    # Read the existing data from the JSON file (if it exists)
                    with open(file_path, 'r') as file:
                        data = json.load(file)
                except (FileNotFoundError, json.JSONDecodeError):
                    # If the file doesn't exist or is empty, create an empty list
                    data = []
                # Append the new dictionary to the list
                data.append(dictionary)
                # Save the updated data back to the JSON file
                with open(file_path, 'w') as file:
                    json.dump(data, file, indent=4)

            # Dictionary to append to the list
            append_dict_to_list_and_save(parsed_data, json_file_path)
        elif response.status_code != 200:
            with open('responseError.txt', 'a') as f:
                f.write(site_url + '\n')
    except:
        with open('Error.txt', 'a') as f:
            f.write(site_url + '\n')
workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

for row_num in range(2, 50):
    url = sheet.cell(row=row_num, column=1).value
    data =get_data(url, row_num)
    if data:
        print(data)
    else:
        print(f"{row_num} {url}")
