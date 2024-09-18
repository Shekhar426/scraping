import requests
from bs4 import BeautifulSoup
import openpyxl

# Function to get data using requests and BeautifulSoup
def get_data(site_url, row_num):
    try:
        response = requests.get(site_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            first_ul = soup.find('ul', {'role': 'presentation'})
            if first_ul:
                total_variation_type = first_ul.find_all('li', recursive=False)
            else:
                total_variation_type = []
            if len(total_variation_type) == 2:
                pid = soup.select_one("span.productsku").text
                try:
                    assert len(soup.select("ul[role='presentation']:nth-of-type(1) li .value ul")) > 1
                    optclr1 = soup.select("ul[role='presentation']:nth-of-type(1) li:nth-of-type(1) .value a")
                    optsize2 = soup.select("ul[role='presentation']:nth-of-type(1) li:nth-of-type(2) .value a")
                    baseUrl = "https://www.scheels.com/on/demandware.store/Sites-scheels-Site/en_US/Product-Variation?"
                    productUrlsoptclr_List = []
                    productUrlsoptsize_List = []
                    for optclr in optclr1:
                        for optsize in optsize2:
                            productUrlsoptclr = optclr['data-href'].split('?')[1].split('&')
                            for puoc in productUrlsoptclr:
                                if "pid=" not in puoc and puoc not in productUrlsoptclr_List:
                                    productUrlsoptclr_List.append(puoc)
                            productUrlsoptsize = optsize['data-href'].split('?')[1].split('&')
                            for puos in productUrlsoptsize:
                                if "pid=" not in puos and puos not in productUrlsoptsize_List:
                                    productUrlsoptsize_List.append(puos)

                    if len(productUrlsoptclr_List) > 1 and len(productUrlsoptsize_List) > 1:
                        for clrs in productUrlsoptclr_List:
                            for sizes in productUrlsoptsize_List:
                                    customizedUrl = f"{baseUrl}pid={pid}&{clrs}&{sizes}"
                                    with open('productUrls.txt', 'a') as f5:
                                        f5.write(customizedUrl + '\n')
                        with open('success.txt', 'a') as f13:
                            f13.write(site_url + '\n')
                    elif len(productUrlsoptclr_List) == 1 and len(productUrlsoptsize_List) > 1:
                        for sizefetchUrl in optsize2:
                            sizeUrls = sizefetchUrl.get('data-href')
                            with open('productUrls.txt', 'a') as f6:
                                f6.write(sizeUrls + '\n')
                        with open('success.txt', 'a') as f14:
                            f14.write(site_url + '\n')
                    elif len(productUrlsoptclr_List) == 1 and len(productUrlsoptsize_List) == 1:
                        print("Waiting logic to be implemented")
                    Options = f"{len(productUrlsoptclr_List) * len(productUrlsoptsize_List)}"
                    return f"{row_num} {site_url} Variation : {len(total_variation_type)} Options : {Options}"
                except:
                    return None
        elif response.status_code != 200:
            with open('responseError.txt', 'a') as f:
                f.write(site_url + '\n')
    except:
        with open('Error.txt', 'a') as f:
            f.write(site_url + '\n')
workbook = openpyxl.load_workbook("../Pre_Processing/urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

for row_num in range(2000, 4000):
    url = sheet.cell(row=row_num, column=1).value
    data =get_data(url, row_num)
    if data:
        print(data)
    else:
        print(f"{row_num} {url}")
