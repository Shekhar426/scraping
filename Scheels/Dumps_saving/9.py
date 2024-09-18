import requests
import openpyxl
from unidecode import unidecode

def get_data(site_url, row_num):
    try:
        html_path = 'E:\\dumps_scheels\\OneGrid\\'
        response = requests.get(site_url)
        response.raise_for_status()  # Check for HTTP errors
        html = response.text
        with open(f'{html_path}{row_num}.html', 'w', encoding='utf-8') as f:
            f.write(unidecode(html))
    except requests.RequestException as e:
        error_message = str(e).split('\n')[0]
        print(f"Error: {error_message} {site_url}")
        with open('error_urls_OneGrid.txt', 'a') as f:
            f.write(f"{site_url}\n")

workbook = openpyxl.load_workbook("productUrls.xlsx")
sheet = workbook['OneGrid']
max_row = sheet.max_row


for row_num in range(24000, 27000):
    url = sheet.cell(row=row_num, column=1).value
    if url:  # Ensure URL is not None or empty
        get_data(url, row_num)
        print(row_num, url)
