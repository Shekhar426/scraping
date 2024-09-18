from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from seleniumbase import Driver
from selenium.common.exceptions import NoSuchElementException
from unidecode import unidecode
import pending

driver = Driver(uc=True, headed=False, browser='chrome', headless=True)

wait = WebDriverWait(driver, 1)
def popUpClosed():
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@data-test-id='dismiss-modal']/span"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='no-thanks']/a"))).click()
    except:
        pass
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='no-thanks']/a"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@data-test-id='dismiss-modal']/span"))).click()
    except:
        pass
def get_data(site_url):
    try:
        html_path = 'E:\\dumps_westelm2/'
        driver.get(site_url)
        html = driver.page_source
        with open(f'{html_path}{row_num}.html', 'w', encoding='utf-8') as f:
            f.write(unidecode(html))
    except (Exception, NoSuchElementException) as e:
        error_message = str(e).split('\n')[0]
        print(f"Error: {error_message} {site_url}")
        with open('error_urls.txt', 'a') as f:
            f.write(f"{site_url}\n")

workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet2']
max_row = sheet.max_row

driver.get('https://www.westelm.com/')
popUpClosed()

for row_num in range(70000, 80000):
    if row_num in pending.pendingList:
        url = sheet.cell(row=row_num, column=1).value
        get_data(url)
        print(row_num, url)