import json
import time
from datetime import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC

executable_path = Service("../chromedriver.exe")
chrome_options = Options()

driver = webdriver.Edge()
wait = WebDriverWait(driver, 10)

#Open Xls File which contains urls.
workbook = openpyxl.load_workbook("grainger_urls.xlsx")
sheet = workbook['Sheet']
max_row = sheet.max_row

for row_num in range(1,max_row+1):
    url = sheet.cell(row=row_num, column=1).value  # Column A is index 1
    driver.get(url)
    try:
        driver.find_element(By.XPATH, "//a[@class='close-modal-icon postalCodeCloseBtn']").click()
    except:
        pass
    item_name = driver.find_element(By.XPATH, "//h1[@class='productName desktopPrdTitle']").text
    brand = driver.find_element(By.XPATH, "(//a[@itemprop='Brand'])[2]")
    brand = driver.execute_script("return arguments[0].textContent;", brand)
    try:
        driver.find_element(By.XPATH, "//p[contains(text(),'This product has been DISCONTINUED')]")
        product_status = "Out of Stock"
        price = ''
        price_unit = ''
    except:
        Complete_price = driver.find_element(By.XPATH, "//span[@itemprop='price' and @class = 'price']").text
        product_status = "In Stock"
        price_unit_list = Complete_price.split("/")
        price = price_unit_list[0].replace(" ", "")
        price_unit = price_unit_list[1].replace(" ", "")

    item_Number = driver.find_element(By.XPATH, "//span[@itemprop='productID']").text
    shipping_weight = driver.find_element(By.XPATH, "//li[@id='shippingWeight']").text

    mfr_model = driver.find_element(By.XPATH, "(//span[@itemprop='model'])[1]").text
    unspsc = driver.find_element(By.XPATH, "(//span[@itemprop='model'])[2]").text
    catalogue_page = driver.find_element(By.XPATH, "//li[@class='catalogPageNum']").text

    category = driver.find_element(By.XPATH, "(//a[@class='bread-link'])[1]")
    category = driver.execute_script("return arguments[0].textContent;", category)
    sub_Category = driver.find_element(By.XPATH, "(//a[@class='bread-link'])[2]")
    sub_Category = driver.execute_script("return arguments[0].textContent;", sub_Category)
    department = driver.find_element(By.XPATH, "(//a[@class='bread-link'])[3]")
    department = driver.execute_script("return arguments[0].textContent;", department)
    # Code to fetch Technocal skills.

    technical_specification = {}

    # First Column Technical specifications
    try:
        liCount_First = driver.find_elements(By.XPATH, "//ul[@class='column firstCol']/li")
        for i in range (len(liCount_First)):
            key_el = driver.find_element(By.XPATH, "(//ul[@class='column firstCol']/li)["+str(i+1)+"]/span[1]")
            key = driver.execute_script("return arguments[0].textContent;", key_el)
            value_el = driver.find_element(By.XPATH, "(//ul[@class='column firstCol']/li)["+str(i+1)+"]/span[2]")
            value = driver.execute_script("return arguments[0].textContent;", value_el)
            technical_specification[key] = value
    except:
        pass
    # Second Column Technical specifications
    try:
        liCount_Second = driver.find_elements(By.XPATH, "//ul[@class='column']/li")
        for i in range (len(liCount_Second)):
            key_el = driver.find_element(By.XPATH, "(//ul[@class='column']/li)["+str(i+1)+"]/span[1]")
            key = driver.execute_script("return arguments[0].textContent;", key_el)
            value_el = driver.find_element(By.XPATH, "(//ul[@class='column']/li)["+str(i+1)+"]/span[2]")
            value = driver.execute_script("return arguments[0].textContent;", value_el)
            technical_specification[key] = value
    except:
        pass
    parsed_data = {'time_stamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'tenant_id': 'Global Industries', 'competitor_id': '', 'competitor_name': 'grainger', 'upc_code': unspsc, 'sku': item_Number, 'title': item_name, 'url': driver.current_url, 'brand': brand, 'category': category, 'description': '', 'quantity': '', 'size': '', 'weight': shipping_weight.replace('Shipping Weight ', ''), 'regular_price': price, 'discount': '', 'unprocessed_json': {"product_Status": product_status, 'price_unit':price_unit, 'sub_category': sub_Category, 'department': department, 'mfr_model': mfr_model, **technical_specification}}; print(parsed_data)
    json_file_path = 'grainger_data.json'
    # Function to append a dict to a list and save it to a JSON file
    def append_dict_to_list_and_save(dictionary, file_path):
        try:
            # Read the existing data from the JSON file (if it exists)
            with open(file_path, 'r') as file:
                data = json.load(file)
        except (FileNotFoundError, json.JSONDecodeError):
            # If the file doesn't exist or is empty, create an empty list
            data = []
        data.append(dictionary)
        # Save the updated data back to the JSON file
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=4)
    # Dictionary to append to the list
    append_dict_to_list_and_save(parsed_data, json_file_path)
