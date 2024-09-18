from selenium.webdriver.support.ui import WebDriverWait
from sbvirtualdisplay.unicodeutil import unidecode
import openpyxl
from seleniumbase import Driver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from unidecode import unidecode
from pending import pendingList
driver = Driver(browser='Chrome')
wait = WebDriverWait(driver, 30)

def get_data(site_url):
    html_path = 'E:/dumps_mackspw/'
    driver.get(site_url)
    wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//div[@class='header-message']/span[contains(text(), 'Need')]")))
    html = driver.page_source
    with open(f'{html_path}{row_num}.html', 'w', encoding='utf-8') as f:
        f.write(unidecode(html))


workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row


for row_num in range(900, 1200):
        url = sheet.cell(row=row_num, column=1).value
        try:
            get_data(url)
            print(row_num, url)
        except:
            with open('error_urls.txt', 'a') as f:
                f.write(f"{url}\n")