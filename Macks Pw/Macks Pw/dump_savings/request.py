import json

from selenium.webdriver.support.ui import WebDriverWait
from sbvirtualdisplay.unicodeutil import unidecode
import openpyxl
from seleniumbase import Driver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from unidecode import unidecode
from pending import pendingList
driver = Driver(browser='Chrome')
wait = WebDriverWait(driver, 30)

workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

for row_num in range(5500, max_row+1):
        url = sheet.cell(row=row_num, column=1).value
        driver.get(url)
        print(url)

        xpath = "(//script[@type='application/ld+json' and contains(text(), 'Product') and contains(text(), 'sku')])[1]"

        element = wait.until(EC.presence_of_element_located((By.XPATH, xpath))).click()
        inner_html = element.get_attribute("innerHTML")
        # Load the inner HTML content into a JSON object
        json_data = json.loads(inner_html)
        print(json_data)
        print(type(json_data))


