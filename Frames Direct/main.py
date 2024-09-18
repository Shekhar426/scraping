import openpyxl
from seleniumbase.common.exceptions import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import json
from datetime import datetime
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By

#Open Xls File which contains urls.
workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Path']
max_row = sheet.max_row

chrome_options = Options()
chrome_options.add_argument("--disable-images")  # Disable image loading
chrome_options.add_argument("--disable-javascript")  # Disable JavaScript
chrome_options.add_argument("--disable-gpu")  # Disable GPU acceleration
chrome_options.add_argument("--disable-software-rasterizer")  # Disable software rasterizer
chrome_options.add_argument("--disable-web-security")  # Disable web security
chrome_options.add_argument("--disable-css")  # Disable CSS
chrome_options.add_argument("--disable-javascript")  # Disable JavaScript
chrome_options.add_argument("--disable-css")  # Disable CSS
chrome_options.add_argument("--blink-settings=imagesEnabled=false")

driver = webdriver.Chrome(options=chrome_options)

# Read data from column A (column index 1) line by line

def data_parser(site_url):
    driver.get(site_url)
    try:
        if driver.current_url.count("/") == 4:
            print("This is a product URL", site_url)
        elif "contact" in site_url:
            script_tag = driver.find_element(By.XPATH, "//script[@type='application/ld+json' and contains(.,'sku')]")
            json_object = script_tag.get_attribute('innerHTML')
            product_data = json.loads(json_object, strict=False)
            variant_sku = product_data.get('sku')
            name = product_data.get('name')
            brand = product_data.get('brand').get('name')
            category = driver.find_element(By.XPATH, "//div[@id='breadcrumb-container']/a[2]").text
            description = driver.find_element(By.XPATH, "//div[@id='prod-seo-description']").text
            try:
                Price_1 = driver.find_element(By.ID, "spanRetailPrice").text
                Price_2 = driver.find_element(By.ID, "spanOurPrice").text
            except:
                try:
                    Price_1 = driver.find_element(By.ID, "spanRetailPrice").text
                    Price_2 = ""
                except:
                    try:
                        Price_2 = driver.find_element(By.ID, "spanOurPrice").text
                        Price_1 = ""
                    except:
                        Price_1 = ""
                        Price_2 = ""
            if len(Price_1) < 2 and  len(Price_2) > 2:
                regular_price = Price_2.replace("$", "")
                discounted_price = ""
            elif len(Price_1) > 2 and  len(Price_2) < 2:
                regular_price = Price_1.replace("$", "")
                discounted_price = ""
            elif len(Price_1) > 2 and  len(Price_2) > 2:
                regular_price = Price_1.replace("$", "")
                discounted_price = Price_2.replace("$", "")
            else:
                regular_price = ""
                discounted_price = ""

            if regular_price == "":
                regular_price = None
            if discounted_price == "":
                discounted_price = None


            image = driver.find_element(By.XPATH, "//img[contains(@id, 'ProductLargeImage')]").get_attribute("src")
            priceCurrency = product_data.get('offers').get('priceCurrency')
            unprocessed_json = {}
            otherinfo1 = driver.find_elements(By.XPATH, "//ul[@id='prod-features-cl']/li")
            for i in range(len(otherinfo1)):
                otherinfo1_text = otherinfo1[i].text.split(":")
                unprocessed_json[otherinfo1_text[0]] = otherinfo1_text[1]
            otherinfo2 = driver.find_elements(By.XPATH, "//div[@id='social-media-bar']/following-sibling::div")
            for i in range(len(otherinfo2)):
                otherinfo2_text = otherinfo2[i].text.split(":")
                unprocessed_json[otherinfo2_text[0]] = otherinfo2_text[1]

            unprocessed_json["variant_url"] = site_url
            parsed_data = {
                    'time_stamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'tenant_id': "Cool frames",
                    'competitor_name': "Frames Direct",
                    'sku': variant_sku,
                    'title': name,
                    'competitor_url': "https://www.framesdirect.com",
                    'brand': brand,
                    'category': category,
                    'description': description,
                    'size': "",
                    'regular_price': regular_price,
                    'discounted_price': discounted_price,
                    "image_url": image,
                    "rating": None,
                    'sub_category': "",
                    "priceCurrency": priceCurrency,
                    "unprocessed_json": {**unprocessed_json}
                }
            print(parsed_data)
            json_file_path = 'Frames_Direct.json'
            # Function to append a dict to a list and save it to a JSON file
            def append_dict_to_list_and_save(dictionary, file_path):
                try:
                    # Read the existing data from the JSON file (if it exists)
                    with open(file_path, 'r') as file:
                        data = json.load(file)
                except (FileNotFoundError, json.JSONDecodeError):
                    # If the file doesn't exist or is empty, create an empty list
                    data = []
                data.append(dictionary)
                # Save the updated data back to the JSON file
                with open(file_path, 'w') as file:
                    json.dump(data, file, indent=4)
            # Dictionary to append to the list
            append_dict_to_list_and_save(parsed_data, json_file_path)

        else:
            try:
                frame_size =  driver.find_element(By.XPATH, "//tr[@class='box-select active-size']").text
            except:
                frame_size = ""


            #Other Available Color Code Below
            try:
                other_colors = {}
                other_color = driver.find_elements(By.XPATH, "//select[contains(@name,'ddlFrameColor')]/option")
                for i in range(len(other_color)):
                    other_colors[f"other_color_{i+1}"] = other_color[i].text
            except:
                other_colors = {}

            #lense material code below
            try:
                lense_materials = {}
                lense_material = driver.find_elements(By.XPATH, "//select[contains(@name, 'lLensMaterial')]/option")
                for i in range(len(lense_material)):
                    lense_materials[f"lense_material_{i+1}"] = lense_material[i].text
            except:
                lense_materials = {}
            #lense type code below
            try:
                lense_types = {}
                lense_type = driver.find_elements(By.XPATH, "//div[@id='rx-lens-type']/span/label[2]")
                for i in range(len(lense_type)):
                    lense_types[f"lense_type_{i+1}"] = lense_type[i].text
            except:
                lense_types = {}
            script_tag = driver.find_element(By.XPATH, "//script[@type='application/ld+json' and contains(.,'sku')]")
            json_object = script_tag.get_attribute('innerHTML')
            product_data = json.loads(json_object, strict=False)

            for i in range((len(product_data.get('offers', {}).get('offers')))):
                try:
                    Price_1 = driver.find_element(By.ID, "spanRetailPrice").text
                    Price_2 = driver.find_element(By.ID, "spanOurPrice").text
                except:
                    try:
                        Price_1 = driver.find_element(By.ID, "spanRetailPrice").text
                        Price_2 = ""
                    except:
                        try:
                            Price_2 = driver.find_element(By.ID, "spanOurPrice").text
                            Price_1 = ""
                        except:
                            Price_1 = ""
                            Price_2 = ""
                if len(Price_1) < 2 and  len(Price_2) > 2:
                    regular_price = Price_2.replace("$", "")
                    discounted_price = ""
                elif len(Price_1) > 2 and  len(Price_2) < 2:
                    regular_price = Price_1.replace("$", "")
                    discounted_price = ""
                elif len(Price_1) > 2 and  len(Price_2) > 2:
                    regular_price = Price_1.replace("$", "")
                    discounted_price = Price_2.replace("$", "")
                else:
                    regular_price = ""
                    discounted_price = ""

                variant_url = None
                color = None
                name = None
                variant_sku = None
                brand = None
                category = None
                description = None
                size = None
                image = None
                ratings = ''
                sub_category = None
                priceCurrency = None
                extra_data = None
                try:
                    variant_url = (product_data.get('offers', {}).get('offers', []))[i].get('url')
                    name = (product_data.get('offers', {}).get('offers', []))[i].get('itemOffered', {}).get('name')
                    variant_sku = (product_data.get('offers', {}).get('offers', []))[i].get('itemOffered', {}).get('sku')
                    brand = (product_data.get('offers', {}).get('offers', []))[i].get('itemOffered', {}).get('brand', {}).get('name')
                    description = (product_data.get('offers', {}).get('offers', []))[i].get('itemOffered', {}).get('description')
                    color = (product_data.get('offers', {}).get('offers', []))[i].get('itemOffered', {}).get('color', {})
                    size = (product_data.get('offers', {}).get('offers', []))[i].get('itemOffered', {}).get('size', {}).get('name')
                    image = (product_data.get('offers', {}).get('offers', []))[i].get('itemOffered', {}).get('image')
                    ratings = product_data.get('aggregateRating', {}).get('ratingValue', None)
                    priceCurrency = (product_data.get('offers', {}).get('offers', []))[i].get('priceCurrency')
                except (AttributeError, IndexError, NoSuchElementException) as e:
                    # Handle the specific exceptions you want to catch
                    print(f"Exception occurred: {e}")
                    # You can choose to log the exception or take other actions as needed

            # Continue with the rest of your code using the handled values or defaults
                extra_data = driver.find_elements(By.XPATH, "//span[@class='prod-desc-items']/..")
                unprocessed_json =  { "size_UI" : frame_size, "other_available_colors": other_colors,"variant_url": variant_url, "lense_material": lense_materials, "lense_type": lense_types}
                for i in range(len(extra_data)):
                    list = extra_data[i].text.split(":")
                    if len(list) > 1:
                        unprocessed_json[list[0].replace("\n", " ")] = list[1]
                    else:
                        pass
                try:
                    category = driver.find_element(By.XPATH, "//div[@id = 'breadcrumb-container']/a[2]").text
                except:
                    category = None
                try:
                    sub_category = driver.find_element(By.XPATH, "//div[@id = 'breadcrumb-container']/a[3]").text
                except:
                    pass
                parsed_data = {
                    'time_stamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'tenant_id': "Cool frames",
                    'competitor_name': "Frames Direct",
                    'sku': variant_sku,
                    'title': name,
                    'competitor_url': "https://www.framesdirect.com",
                    'brand': brand,
                    'category': category,
                    'description': description,
                    'size_json': size,
                    'regular_price': regular_price,
                    'discounted_price': discounted_price,
                    "image_url": image,
                    "rating": ratings,
                    'sub_category': sub_category,
                    "priceCurrency": priceCurrency,
                    "color": color,
                    "unprocessed_json": {**unprocessed_json}
                }
                print(parsed_data)
                json_file_path = 'Frames_Direct.json'
            # Function to append a dict to a list and save it to a JSON file
                def append_dict_to_list_and_save(dictionary, file_path):
                    try:
                        # Read the existing data from the JSON file (if it exists)
                        with open(file_path, 'r') as file:
                            data = json.load(file)
                    except (FileNotFoundError, json.JSONDecodeError):
                        # If the file doesn't exist or is empty, create an empty list
                        data = []
                    data.append(dictionary)
                    # Save the updated data back to the JSON file
                    with open(file_path, 'w') as file:
                        json.dump(data, file, indent=4)
                # Dictionary to append to the list
                append_dict_to_list_and_save(parsed_data, json_file_path)
    except:
        print("Error in URL")
for row_num in range(1, max_row + 1):
    url = sheet.cell(row=row_num, column=1).value  # Column A is index 1
    print(f"{row_num} : {url}")
    data_parser(url)


