import openpyxl
import os
import argparse
import random
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from seleniumbase import Driver
from unidecode import unidecode
import time

# Initialize Driver with headless option
driver = Driver(uc=True, headed=False, browser='chrome', headless=True)
wait = WebDriverWait(driver, 10)

# Argument parser to take range values from command line
# Open Xls File which contains URLs.

workbook = openpyxl.load_workbook("urlsRH.xlsx")
sheet = workbook['Sheet2']
max_row = sheet.max_row

for i in range(1, max_row+1):
    url = sheet.cell(row=1, column=1).value
    driver.get(url)
