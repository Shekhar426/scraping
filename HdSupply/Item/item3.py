import json
import time
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from seleniumbase import Driver

driver = Driver(uc=True, headed=False, browser='chrome', headless=True)
wait = WebDriverWait(driver, 20)

workbook = openpyxl.load_workbook("data.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row


def append_dict_to_list_and_save(dictionary, file_path):
    try:
        # Read the existing data from the JSON file (if it exists)
        with open(file_path, 'r') as file:
            data = json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        # If the file doesn't exist or is empty, create an empty list
        data = []
    # Append the new dictionary to the list
    data.append(dictionary)
    # Save the updated data back to the JSON file
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)


# Dictionary to append to the list

File = "Item_3"
json_file_path = f'{File}.json'

for row_num in range(4400, max_row + 1):
    COMP_NAME = "HDSupply-US"
    LEGACY_NUMBER = sheet.cell(row=row_num, column=1).value
    ITEM_DESC = sheet.cell(row=row_num, column=2).value
    GIVEN_COMP_URL = sheet.cell(row=row_num, column=3).value

    try:
        custom_search_url = f"https://hdsupplysolutions.com/SearchDisplay?categoryId=&storeId=10051&catalogId=10054&langId=-1&sType=SimpleSearch&resultCatEntryType=2&showResultsPage=true&searchSource=Q&pageView=grid&beginIndex=0&pageSize=24&searchTerm={LEGACY_NUMBER}"
        driver.get(custom_search_url)
        try:
            NotFoundErrorCheck = wait.until(EC.visibility_of_all_elements_located((By.XPATH,
                                                                                   "//p[contains(text(), 'Check your spelling or use different keywords and try again')] | //span[contains(text(),'t find any results for')] | //div[contains(text(), 'is no longer available, but you may find similar parts below')] | //input[@class='jq-compare-item']")))
            No_Product_Found = []
            parsed_data = {
                "COMP_NAME": COMP_NAME,
                "LEGACY_NUMBER": LEGACY_NUMBER,
                "ITEM_DESC": ITEM_DESC,
                "GIVEN_COMP_URL": GIVEN_COMP_URL,
                "custom_search_url": custom_search_url,
                "Product_URL": None,
                "Part_number": None
            }
            append_dict_to_list_and_save(parsed_data, json_file_path)
            print(f"{row_num} : {parsed_data}")
        except:
            product_url = wait.until(
                EC.presence_of_element_located((By.XPATH, "//link[@id='canonicalSeoURL']"))).get_attribute('href')
            part_number = wait.until(EC.visibility_of_element_located(
                (By.XPATH, "(//div[@class='pdp-product-info-left pdp-part-number-info']/span)[last()]"))).text
            parsed_data = {
                "COMP_NAME": COMP_NAME,
                "LEGACY_NUMBER": LEGACY_NUMBER,
                "ITEM_DESC": ITEM_DESC,
                "GIVEN_COMP_URL": GIVEN_COMP_URL,
                "custom_search_url": custom_search_url,
                "Product_URL": product_url,
                "Part_number": part_number
            }
            print(f"{row_num} : {parsed_data}")
            append_dict_to_list_and_save(parsed_data, json_file_path)
    except:
        with open(f'{File}.txt', 'a') as file:
            file.write(f"{row_num} : {LEGACY_NUMBER}\n")
            time.sleep(180)
    if row_num%100 == 0:
        time.sleep(180)