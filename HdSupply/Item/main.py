import json

from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from seleniumbase import Driver

driver = Driver(uc=True, headed=False, browser='chrome', headless=False)
wait = WebDriverWait(driver, 20)

workbook = openpyxl.load_workbook("data.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

def append_dict_to_list_and_save(dictionary, file_path):
    try:
        # Read the existing data from the JSON file (if it exists)
        with open(file_path, 'r') as file:
            data = json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        # If the file doesn't exist or is empty, create an empty list
        data = []
    # Append the new dictionary to the list
    data.append(dictionary)
    # Save the updated data back to the JSON file
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)

# Dictionary to append to the list
json_file_path = 'dataCheck .json'


for row_num in range(2, max_row+1):
    url = sheet.cell(row=row_num, column=4).value
    updatedUrl = "https://hdsupplysolutions.ca/"+url.split("https://hdsupplysolutions.com/")[1]
    driver.get(updatedUrl)
    try:
        pageNotFoundError1 = driver.find_element(By.XPATH, "//p[contains(text(), 'An error has occurred.')]").text
        pageNotFoundError2 = driver.find_element(By.XPATH, "//p[contains(text(), 'Please try again in a few minutes or contact us at CustomerCare_CAN@hdsupply.com')]").text
        parsed_data = {
            "row_num": row_num,
            "url": url,
            "updatedUrl": updatedUrl,
            "error": "Yes"
        }
        append_dict_to_list_and_save(parsed_data, json_file_path)
    except:
        parsed_data = {
            "row_num": row_num,
            "url": url,
            "updatedUrl": updatedUrl,
            "error": "No"
        }
        append_dict_to_list_and_save(parsed_data, json_file_path)
    print(row_num)

