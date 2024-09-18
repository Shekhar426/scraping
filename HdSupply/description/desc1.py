import json
import time

from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from seleniumbase import Driver

driver = Driver(uc=True, headed=False, browser='chrome', headless=False)
wait = WebDriverWait(driver, 20)

workbook = openpyxl.load_workbook("data.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

def append_dict_to_list_and_save(dictionary, file_path):
    try:
        # Read the existing data from the JSON file (if it exists)
        with open(file_path, 'r') as file:
            data = json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        # If the file doesn't exist or is empty, create an empty list
        data = []
    # Append the new dictionary to the list
    data.append(dictionary)
    # Save the updated data back to the JSON file
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)
# Dictionary to append to the list

File = "Desc_1"
json_file_path = f'{File}.json'

for row_num in range(2, 2200):
    COMP_NAME = "HDSupply-US"
    LEGACY_NUMBER = sheet.cell(row=row_num, column=1).value
    ITEM_DESC = sheet.cell(row=row_num, column=2).value
    GIVEN_COMP_URL = sheet.cell(row=row_num, column=3).value
    try:
        driver.get("https://hdsupplysolutions.com")
        wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@placeholder='Search keyword or part number']"))).send_keys(ITEM_DESC)
        wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@placeholder='Search keyword or part number']"))).send_keys(Keys.ENTER)

        allProducts = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//div[@class='subcat-grid-tile__description']/a")))
        UrlsList = []
        for url in allProducts[0:10]:
            UrlsList.append(url.get_attribute('href'))
        custom_search_url = driver.current_url
        parsed_data = {
            "COMP_NAME": COMP_NAME,
            "LEGACY_NUMBER": LEGACY_NUMBER,
            "ITEM_DESC": ITEM_DESC,
            "GIVEN_COMP_URL": GIVEN_COMP_URL,
            "custom_search_url": custom_search_url,
            "Product_URL": UrlsList,
        }
        print(parsed_data)
        append_dict_to_list_and_save(parsed_data, json_file_path)
    except:
        with open(f'{File}.txt', 'a') as file:
            file.write(f"{row_num} : {LEGACY_NUMBER}\n")