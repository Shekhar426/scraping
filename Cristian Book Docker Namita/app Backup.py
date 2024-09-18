
import openpyxl
import os
import argparse
import random

from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from seleniumbase import Driver
from unidecode import unidecode
import time
import completed

# Argument parser to take range values from command line
parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument('--start', type=int, required=True, help='Starting row number')
parser.add_argument('--end', type=int, required=True, help='Ending row number')
args = parser.parse_args()

# Open Xls File which contains URLs.
workbook = openpyxl.load_workbook("urlsRH.xlsx")
sheet = workbook['Sheet2']

def create_folder_if_not_exists(folder_path):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"Folder created at: {folder_path}", end=" | ")
    else:
        print(f"Folder already exists at: {folder_path}")
def dumps_downloading(i):
    count = 0
    path = f'/usr/src/app/dumps_rh/{i}/'
    driver.get(url)
    wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='MuiGrid-root MuiGrid-container']//h1")))
    print(driver.title)
    Flag = False
    if "Access" in driver.title:
        Flag = True
    tryCount = 0
    while Flag == True:
        time.sleep(random.choice([10,15,20,25,26,27,29,30]))
        driver.reload()
        tryCount = tryCount+1
        if "Access" not in driver.title:
            print("Access Denied Error is resolved  {}".format(i))
            Flag = False
            break
        elif tryCount > 10:
            print("Max Try for Access Denied Moving to next url  {}".format(i))
            Flag = False
            break
        print("Tried Count  {}. Please wait let's try few more time........".format(tryCount))
    wait.until(EC.visibility_of_element_located((By.XPATH, "//h1[contains(@class, 'MuiTypography-h1')]")))

    # Check that page is related to first condition. if page condition is of first type then set variable value as First
    try:
        arrowPointer = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//div[contains(@class, 'flex items-center justify-between pr-2 relative') and contains(@data-testid, 'SELECT')]")))
    except:
        arrowPointer = []
    try:
        typeOfGrid = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//div[@id='component-swatch-groups']")))
    except:
        typeOfGrid = []

    # If there are only colors and no arrow pointer available to click

    if len(arrowPointer) == 0 and len(typeOfGrid) == 1:
        combinationElement = wait.until(
            EC.visibility_of_all_elements_located((By.XPATH, "//div[@id='component-swatch-groups']//ul/li")))
        optionsName = wait.until(EC.visibility_of_all_elements_located(
            (By.XPATH, "//div[@id='component-swatch-groups']//ul/li//p[not(contains(text(), 'Sale'))]")))
        print("Total options found", len(combinationElement), len(optionsName))
        create_folder_if_not_exists(path)
        if len(combinationElement) == len(optionsName):
            dumpCountFor_PerticularURL = 0
            print("Total combination found", len(combinationElement))
            for index in range(len(combinationElement)):
                combinationElement[index].click()
                html = driver.page_source
                with open('{}{}_{}.html'.format(path, sequence,optionsName[index].text.replace("/", "_").replace("\n", "_")), 'w',
                          encoding='utf-8') as f:
                    f.write(unidecode(html))
                print("Html dump saved for combination")
                print(f"URL: {url} | Sequence: {sequence} | Count: {count} | Action: Dumps Data Saved. Total Dumps Saved: {dumpCountFor_PerticularURL}")
                dumpCountFor_PerticularURL += 1
            count += 1
            with open(f'rhSuccess.txt', 'a') as f:
                f.write(url)
                f.write('\n')
        else:
            with open(f'rhFailure.txt', 'a') as f3:
                f3.write(url)
                f3.write('\n')
            count += 1
            assert 1 == 2
    elif len(arrowPointer) == 1 and len(typeOfGrid) == 1:
        time.sleep(3)
        wait.until(EC.visibility_of_element_located((By.XPATH, '//div[contains(@data-testid,"SELECT FROM")]'))).click()
        time.sleep(3)
        checkColorLength = wait.until(EC.visibility_of_all_elements_located((By.XPATH, '//h3[contains(@class,"MuiTypography-h3")]/span/../../following-sibling::div//ul/li')))
        loopCount = len(checkColorLength)
        checkColorLength[0].send_keys(Keys.ESCAPE)
        for k in range(loopCount):
            time.sleep(3)
            element = driver.find_element(By.XPATH, "//a[@id='container-rhr-header_logo-rhr']")
            action = ActionChains(driver)
            action.move_to_element(element).perform()
            time.sleep(3)
            driver.find_element(By.XPATH, '//div[contains(@data-testid,"SELECT FROM")]').click()
            time.sleep(3)
            checkColorLength = wait.until(EC.visibility_of_all_elements_located((By.XPATH, '//h3[contains(@class,"MuiTypography-h3")]/span/../../following-sibling::div//ul/li')))
            time.sleep(3)
            allColorName = wait.until(EC.visibility_of_all_elements_located((By.XPATH, '//h3[contains(@class,"MuiTypography-h3")]/span/../../following-sibling::div//ul/li')))
            colorName = allColorName[k].text
            time.sleep(3)
            checkColorLength[k].click()
            checkColorLength[k].send_keys(Keys.ESCAPE)
            time.sleep(3)
            combinationElement = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//div[@id='component-swatch-groups']//ul/li")))
            time.sleep(3)
            optionsName = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//div[@id='component-swatch-groups']//ul/li//p[not(contains(text(), 'Sale'))]")))
            create_folder_if_not_exists(path)
            if len(combinationElement) == len(optionsName):
                dumpCountFor_PerticularURL = 0
                print("Total combination found", len(combinationElement))
                for index in range(len(combinationElement)):
                    time.sleep(3)
                    combinationElement[index].click()
                    html = driver.page_source
                    with open('{}{}_{}_{}_{}.html'.format(path, sequence, colorName, k,optionsName[index].text.replace("/", "_").replace("\n", "_")),
                              'w', encoding='utf-8') as f:
                        f.write(unidecode(html))
                    print(f"URL: {url} | Sequence: {sequence} | Count: {count} | Action: Dumps Data Saved. Total Dumps Saved: {dumpCountFor_PerticularURL}")
                    dumpCountFor_PerticularURL += 1
                count += 1
                with open(f'rhSuccess.txt', 'a') as f:
                    f.write(url)
                    f.write('\n')
        count += 1

    elif len(arrowPointer) == 1 and len(typeOfGrid) == 0: # Third condition where there is one arrow pointer but no grids found
        time.sleep(4)
        wait.until(EC.visibility_of_element_located((By.XPATH, '//div[contains(@data-testid,"SELECT FROM")]'))).click()
        time.sleep(4)
        checkColorLength = wait.until(EC.visibility_of_all_elements_located((By.XPATH, '//h3[contains(@class,"MuiTypography-h3")]/span/../../following-sibling::div//ul/li')))
        loopCount = len(checkColorLength)
        checkColorLength[0].send_keys(Keys.ESCAPE)
        time.sleep(4)
        for k in range(loopCount):
            time.sleep(4)
            driver.find_element(By.XPATH, '//div[contains(@data-testid,"SELECT FROM")]').click()
            time.sleep(4)
            checkColorLength = wait.until(EC.visibility_of_all_elements_located((By.XPATH, '//h3[contains(@class,"MuiTypography-h3")]/span/../../following-sibling::div//ul/li')))
            time.sleep(4)
            allColorName = wait.until(EC.visibility_of_all_elements_located((By.XPATH, '//h3[contains(@class,"MuiTypography-h3")]/span/../../following-sibling::div//ul/li')))
            colorName = allColorName[k].text
            time.sleep(4)
            checkColorLength[k].click()
            checkColorLength[k].send_keys(Keys.ESCAPE)
            time.sleep(4)
            create_folder_if_not_exists(path)
            dumpCountFor_PerticularURL = 0
            html = driver.page_source
            with open('{}{}_{}_{}.html'.format(path, sequence, colorName, k), 'w', encoding='utf-8') as f:
                f.write(unidecode(html))
            print(f"URL: {url} | Sequence: {sequence} | Count: {count} | Action: Dumps Data Saved. Total Dumps Saved: {dumpCountFor_PerticularURL}")
            dumpCountFor_PerticularURL += 1
        count += 1
        with open(f'rhSuccess.txt', 'a') as f:
            f.write(url)
            f.write('\n')
        count += 1

    elif len(arrowPointer) == 0 and len(typeOfGrid) == 0:
        html = driver.page_source
        with open('{}{}.html'.format(path, sequence), 'w', encoding='utf-8') as f:
            f.write(unidecode(html))
        print(f"URL: {url} | Sequence: {sequence} | Count: {count} | Action: Dumps Data Saved")
        count += 1
        with open(f'rhSuccess.txt', 'a') as f:
            f.write(url)
            f.write('\n')
# Initialize Driver with headless option
driver = Driver(uc=True, headed=False, browser='chrome', headless=True)
wait = WebDriverWait(driver, 10)

for i in range(args.start, args.end + 1):
    if i not in completed.completed:
        sequence = sheet.cell(row=i, column=1).value
        url = sheet.cell(row=i, column=2).value
        time.sleep(random.choice([0, 1, 2,0, 2,5,6,2,7,9,3,0]))
        print("Row Number : {} {}".format(i, url), end = " | ") 
        try:
            dumps_downloading(i)
        except Exception as e:
            print(e)
            with open(f'/usr/src/app/dumps_rh/rhFailure.txt', 'a') as f3:
                f3.write(url)
                f3.write('\n')