import argparse
import json

from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from seleniumbase import Driver

driver = Driver(browser='Chrome')
wait = WebDriverWait(driver, 20)

# Argument parser to take range values from command line
parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument('--start', type=int, required=True, help='Starting row number')
parser.add_argument('--end', type=int, required=True, help='Ending row number')
args = parser.parse_args()

workbook = openpyxl.load_workbook("data.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

def append_dict_to_list_and_save(dictionary, file_path):
    try:
        # Read the existing data from the JSON file (if it exists)
        with open(file_path, 'r') as file:
            data = json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        # If the file doesn't exist or is empty, create an empty list
        data = []
    # Append the new dictionary to the list
    data.append(dictionary)
    # Save the updated data back to the JSON file
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)


# Dictionary to append to the list


json_file_path = f'/usr/src/app/Book/{args.start}.json'

for row_num in range(args.start, args.end):

    ITEM = sheet.cell(row=row_num, column=1).value
    ITEM_NAME_DESCRIPTION = sheet.cell(row=row_num, column=2).value
    PRODUCT_TYPE = sheet.cell(row=row_num, column=3).value
    UPC = sheet.cell(row=row_num, column=4).value
    ISBN13 = sheet.cell(row=row_num, column=5).value
    ISBN = sheet.cell(row=row_num, column=6).value
    PRODUCT = sheet.cell(row=row_num, column=7).value

    driver.get("https://www.christianbook.com")

    driver.find_element(By.XPATH, "//input[@id='Ntt']").send_keys(ITEM)
    driver.find_element(By.XPATH, "//input[@id='Ntt']").send_keys(Keys.ENTER)

    print("Identifying Page Type...")

    try:
        Product_Name = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//h1[@class='CBD-ProductDetailTitle']")))
        No_Product_Found = []
        Multi_product_Found = []
    except:
        try:
            No_Product_Found = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//h1[text()='No Search Results Found']")))
            Product_Name = []
            Multi_product_Found = []
        except:
            try:
                Multi_product_Found = wait.until(EC.visibility_of_all_elements_located((By.XPATH, "//p[@class='CB-ProductListItem-Title']")))
                Product_Name = []
                No_Product_Found = []
            except:
                Product_Name = []
                No_Product_Found = []
                Multi_product_Found = []

    print("Page Identification Completed Successfully...")

    if len(Product_Name) == 1 and len(No_Product_Found) == 0 and len(Multi_product_Found) == 0:
        print("Product found. Entered in IF Condition")
        productUrl = driver.find_element(By.XPATH, "//meta[@property='og:url']").get_attribute('content')
        images = []
        allImages = driver.find_elements(By.XPATH, "//div[@id='CBD-ProductThumbnailWrapper']/ul/li/span/..")
        if len(allImages) > 0:
            for img in allImages:
                singleImage1 = img.get_attribute('data-fullsrc')
                if "https" not in singleImage1:
                    singleImage = "https:" + singleImage1
                    images.append(singleImage)
                else:
                    images.append(singleImage1)
        else:
            singleImage2 = driver.find_element(By.XPATH,"//div[@class='CBD-ProductImageContainer']//img").get_attribute('src')
            images.append(singleImage2)
        parsed_data = {
            "item": ITEM,
            "item_name_description": ITEM_NAME_DESCRIPTION,
            "product_type": PRODUCT_TYPE,
            "upc": UPC,
            "isbn13": ISBN13,
            "isbn": ISBN,
            "product": PRODUCT,
            "productUrl": productUrl,
            "images": images
        }
        append_dict_to_list_and_save(parsed_data, json_file_path)
    elif len(Product_Name) == 0 and len(No_Product_Found) == 1 and len(Multi_product_Found) == 0:
        print("No Product found. Entered in First ELIF Condition")
        parsed_data = {
            "item": ITEM,
            "item_name_description": ITEM_NAME_DESCRIPTION,
            "product_type": PRODUCT_TYPE,
            "upc": UPC,
            "isbn13": ISBN13,
            "isbn": ISBN,
            "product": PRODUCT,
            "productUrl": None,
            "images": None
        }
        append_dict_to_list_and_save(parsed_data, json_file_path)
    elif len(Product_Name) == 0 and len(No_Product_Found) == 0 and len(Multi_product_Found) > 0:
        print("Multi Product Found. Entered in Second ELIF Condition")
        try:
            wait.until(EC.element_to_be_clickable((By.XPATH, f"//p[@class='CB-ProductListItem-Title']/ancestor::div[@class='CB-ProductListItem-InfoContent']//p[contains(text(), '{ITEM}')]"))).click()
            productUrl = driver.find_element(By.XPATH, "//meta[@property='og:url']").get_attribute('content')
            images = []
            allImages = driver.find_elements(By.XPATH, "//div[@id='CBD-ProductThumbnailWrapper']/ul/li/span/..")
            if len(allImages) > 0:
                for img in allImages:
                    singleImage1 = img.get_attribute('data-fullsrc')
                    if "https" not in singleImage1:
                        singleImage = "https:" + singleImage1
                        images.append(singleImage)
                    else:
                        images.append(singleImage1)
            else:
                singleImage2 = driver.find_element(By.XPATH,"//div[@class='CBD-ProductImageContainer']//img").get_attribute('src')
                images.append(singleImage2)
            parsed_data = {
                "item": ITEM,
                "item_name_description": ITEM_NAME_DESCRIPTION,
                "product_type": PRODUCT_TYPE,
                "upc": UPC,
                "isbn13": ISBN13,
                "isbn": ISBN,
                "product": PRODUCT,
                "productUrl": productUrl,
                "images": images
            }
            append_dict_to_list_and_save(parsed_data, json_file_path)
        except:
            parsed_data = {
                "item": ITEM,
                "item_name_description": ITEM_NAME_DESCRIPTION,
                "product_type": PRODUCT_TYPE,
                "upc": UPC,
                "isbn13": ISBN13,
                "isbn": ISBN,
                "product": PRODUCT,
                "productUrl": None,
                "images": None
            }
            append_dict_to_list_and_save(parsed_data, json_file_path)
    print(f"Row {row_num} processed")