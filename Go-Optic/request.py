import openpyxl
import json
from datetime import datetime
import requests
from lxml import html

workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row

def get_data(url):
    response = requests.get(url)
    tree = html.fromstring(response.content)
    name = tree.xpath("//div[@class='detail-area wf-primary']/h4/b/text()")[0]
    sku = url.split('/p/')[-1]
    brand = tree.xpath("//div[@class='made-by']/a/text()")[0]
    try:
        regular_price = tree.xpath("//div[@class='original-price']/s/text()")[0].replace("$", "")
    except IndexError:
        regular_price = tree.xpath("//div[@class='price']/span/text()")[0].replace("$", "")
    parsed_data = {
    'time_stamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    'sku': sku,
    'title': name,
    'brand': brand,
    'regular_price': regular_price,
    "variant_url": url,}
    print(parsed_data)
for row_num in range(1, max_row + 1):
    url = sheet.cell(row=row_num, column=1).value
    print(row_num, url)
    get_data(url)
