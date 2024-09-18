import json
from datetime import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

executable_path = Service("../chromedriver.exe")
chrome_options = Options()
#chrome_options.add_argument("--headless")

driver = webdriver.Edge()
wait = WebDriverWait(driver, 3)

#Open Xls File which contains urls.
workbook = openpyxl.load_workbook("Valleyvetdata.xlsx")
sheet = workbook['UniqueProductUrls']
max_row = sheet.max_row

# Read data from column A (column index 1) line by line

for row_num in range(1):
    main_dict = {}
    url = "https://www.valleyvet.com/ct_detail.html?pgguid=e13bc452-0bed-4b00-8674-1b3e48382416"
    # url = sheet.cell(row=row_num, column=1).value  # Column A is index 1
    driver.get(url)
    def get_data():
        try:
            if driver.title != "Horse Supplies, Pet Supplies, Farm Supplies, Goat Supplies - Valley Vet Supply":
                #Close popup if visible on the page
                try:
                    driver.find_element(By.XPATH, "//button[@class='ltkpopup-close']").click()
                except:
                    pass
                # If "Terms of Use" is visible on the page close that.
                try:
                    driver.find_element(By.XPATH, "//div[@id='ToU_float']/a").click()
                except:
                    pass
                tableRow = driver.find_elements(By.XPATH, "//tbody/tr")
                title = driver.find_element(By.XPATH, "//h1").text
                category_text = driver.find_element(By.XPATH, "//nav[@class='breadcrumb']/a").text
                if category_text.startswith("-"):
                    category = ""
                else:
                    category = category_text
                description = driver.find_element(By.XPATH, "//meta[@name='description']").get_attribute("content")
                try:
                    brand = driver.find_element(By.XPATH, "//div[@class='by']/a").text
                except:
                    brand = ""
                column_count = 0
                column = driver.find_elements(By.XPATH, "//table[@id='tableOptions']/thead/tr/th")
                column_list = []
                column_list.append(url)
                table = driver.find_element(By.XPATH, "//table")
                data_dict = {}
                header_cells = table.find_elements(By.XPATH, ".//tr/th")
                column_names = []
                for cell in header_cells:
                    CL = (cell.text.split("\n"))
                    if len(CL) > 1:
                        if "MAP " in CL[1]:
                            data_dict["MAP_STATUS"] = "YES"
                            column_names.append(CL[1].replace("MAP ", ""))
                        else:
                            data_dict["MAP_STATUS"] = "NO"
                            column_names.append(CL[1])
                    else:
                        pass
                # Initialize lists for each column
                for column_name in column_names:
                    data_dict[column_name] = ""
                # Find all the rows in the table
                rows = table.find_elements(By.XPATH, ".//tr")
                # Iterate through the rows to extract data and build the dict
                for row in rows[1:]:  # Skip the header row
                    cells = row.find_elements(By.XPATH, ".//td")
                    for i, cell in enumerate(cells):
                        if i < len(column_names):
                            if cell.text.startswith("$"):
                                if len(cell.text.split(" "))  == 1:
                                    cell_price = cell.text.split(" ")
                                    data_dict[column_names[i]] = (cell_price[0])
                                else:
                                    cell_price = cell.text.split(" ")
                                    data_dict[column_names[i]] = (cell_price[0])
                                    data_dict["DISSCOUNT"] = (cell_price[1])
                            else:
                                data_dict[column_names[i]] = (cell.text)
                        else:
                            continue
                    main_dict["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    main_dict["tenant_id"] = "REVIVALANIMALHEALTH"
                    main_dict["competitor_id"] = ""
                    main_dict["competitor_name"] = "valleyvet"
                    main_dict["upc_code"] = ""
                    main_dict["competitor_name"] = "valleyvet"
                    main_dict["sku"] = data_dict.get("ITEM")
                    main_dict["title"] = title
                    main_dict["url"] = driver.current_url
                    main_dict["brand"] = brand
                    main_dict["category"] = category
                    main_dict["description"] = description
                    main_dict["quantity"] = ""
                    main_dict["size"] = data_dict.get("SIZE")
                    main_dict["weight"] = ""
                    main_dict["regular_price"] = data_dict.get("PRICE")
                    main_dict["discount"] = data_dict.get("DISSCOUNT")
                    main_dict["unprocessed_json"] = data_dict

                    json_file_path = 'data.json'
                    # Function to append a dict to a list and save it to a JSON file
                    def append_dict_to_list_and_save(dictionary, file_path):
                        try:
                            # Read the existing data from the JSON file (if it exists)
                            with open(file_path, 'r') as file:
                                data = json.load(file)
                        except (FileNotFoundError, json.JSONDecodeError):
                            # If the file doesn't exist or is empty, create an empty list
                            data = []
                        data.append(dictionary)
                        # Save the updated data back to the JSON file
                        with open(file_path, 'w') as file:
                            json.dump(data, file, indent=4)
                    # Dictionary to append to the list
                    append_dict_to_list_and_save(main_dict, json_file_path)
                    print(main_dict)
            else:
                print(url, "PRODUCT NOT FOUND")
        except:
            Category_page = driver.find_elements(By.XPATH, "//div[@title='Items per Page']")
            print("{} : Not a product URL".format(url))
            pass
    get_data()
