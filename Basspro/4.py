import openpyxl
import time
import os
import random
from selenium.webdriver.common.keys import Keys
from selenium.common import InvalidSessionIdException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from seleniumbase import Driver
from seleniumbase.common.exceptions import WebDriverException
from unidecode import unidecode
from argparse import ArgumentParser

workbook = openpyxl.load_workbook("urls.xlsx")
sheet = workbook['Sheet3']
max_row = sheet.max_row

def dumps_downloading(i,url):
    try:
        driver.get(url)
        html = driver.page_source
        with open(f"E:\\dumps_basspro/{i}.html", 'w', encoding='utf-8') as f:
                f.write(unidecode(html))
    except:
         with open(f'bassproError.txt', 'a') as f:
                    f.write(url)
                    f.write('\n')

driver=Driver(browser='Chrome')
wait = WebDriverWait(driver, 40)

driver.get("https://www.basspro.com/shop/en")

wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@class="BPSSLCity_State"]'))).click()
wait.until(EC.presence_of_element_located((By.XPATH, "//input[@id='storeLocInp']"))).send_keys("65616")
wait.until(EC.presence_of_element_located((By.XPATH, "//input[@id='storeLocInp']"))).send_keys(Keys.ENTER)
# time.sleep(7) # input Text ZIP Code
wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='store_locator_add' and contains(text(),'65616')]/following-sibling::button"))).click()
time.sleep(3)
for i in range(1440,1460):
    url = sheet.cell(row=i, column=1).value
    dumps_downloading(i,url)
    print(i, url)