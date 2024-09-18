import openpyxl


workbook = openpyxl.load_workbook("data.xlsx")
sheet1 = workbook['Sheet1']
max_row1 = sheet1.max_row


workbook = openpyxl.load_workbook("ImageDownloading/output.xlsx")
sheet2 = workbook['Sheet1']
max_row2 = sheet1.max_row

data = []
for i in range(max_row1+1):
    data1 = sheet1.cell(row=i+1, column=1).value
    data.append(data1)
print(len(data))


output = []
for i in range(max_row2+1):
    data2 = sheet2.cell(row=i+1, column=1).value
    output.append(data2)

missing = []
for item in data:
    if item not in output:
        missing.append(item)

print(missing)
print(len(missing))